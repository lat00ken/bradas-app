import os
os.environ["KIVY_GL_BACKEND"] = "sdl2"
from kivy.config import Config

# Detect Android early (python-for-android sets ANDROID_ARGUMENT)
_IS_ANDROID = "ANDROID_ARGUMENT" in os.environ

# Desktop dev defaults (on Android we let the OS manage the window/surface size)
if not _IS_ANDROID:
    Config.set('graphics', 'width', '1280')
    Config.set('graphics', 'height', '800')
    Config.set('graphics', 'resizable', '0')

Config.set("input", "mouse", "mouse")  # для ПК-отладки (колесо/drag). На планшете ок.
import shutil
import csv
import json
import re
import datetime
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from kivy.app import App
from kivy.lang import Builder
from kivy.clock import Clock
from kivy.core.text import LabelBase
from kivy.logger import Logger
from kivy.properties import StringProperty, NumericProperty, ObjectProperty, ListProperty, BooleanProperty
from kivy.metrics import dp, sp
from kivy.graphics import Color, Rectangle, RoundedRectangle, Line
from kivy.core.window import Window
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.relativelayout import RelativeLayout
from kivy.uix.recycleview import RecycleView
from kivy.uix.recycleview.views import RecycleDataViewBehavior
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserIconView, FileChooserListView
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
class PopupButton(Button):
    """Unified popup button EXACTLY like '+' popup buttons."""
    up_color = ListProperty([0.30, 0.30, 0.30, 1])
    down_color = ListProperty([0.1529, 0.6784, 0.9608, 1])

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ""
        self.background_down = ""
        self.background_color = self.up_color
        self.color = (1, 1, 1, 1)
        self.font_size = dp(18)
        self.size_hint_y = None
        self.height = dp(54)
        self.bind(state=self._sync_state)

    def _sync_state(self, *_):
        self.background_color = self.down_color if self.state == "down" else self.up_color
        self.background_color = self.down_color if self.state == "down" else self.up_color
from kivy.uix.gridlayout import GridLayout
from kivy.uix.widget import Widget
def _try_register_mdl2():
    """Try to register Segoe MDL2 Assets on Windows for icon-like glyphs."""
    try:
        if os.name != "nt":
            return False
        candidates = [
            r"C:\\Windows\\Fonts\\segmdl2.ttf",
            r"C:\\Windows\\Fonts\\SegoeMDL2Assets.ttf",
        ]
        for p in candidates:
            if os.path.exists(p):
                LabelBase.register(name="mdl2", fn_regular=p)
                return True
    except Exception:
        pass
    return False
_MDL2_OK = _try_register_mdl2()
# ---------- account file signature ----------
ACCOUNT_FILE_MAGIC = 'PLANWET_ACCOUNT'
ACCOUNT_FILE_VERSION = 1
ACCOUNT_FILE_MAGIC_VALUE = f"{ACCOUNT_FILE_MAGIC}_V{ACCOUNT_FILE_VERSION}"
# ---------- helpers ----------
def to_float(x):
    if x is None:
        return 0.0
    s = str(x).strip()
    if not s:
        return 0.0
    s = s.replace("\u00a0", " ").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0
def load_products_from_export_csv(csv_path: str):
    """
    export.csv (delimiter ';')
    Беремо:
      - Артикул
      - Назва
      - Ціна ОПТ (USD)         -> закупка (опт)
      - Ціна Роздрібна (USD)   -> роздріб (USD)
    Показуємо ТІЛЬКИ товари, де:
      - Артикул не пустий
      - Ціна ОПТ (USD) > 0
      - Ціна Роздрібна (USD) > 0   (як і домовлялись — приховуємо якщо 0/порожня)
    """
    if not os.path.exists(csv_path):
        return [], f"Файл не знайдено: {os.path.basename(csv_path)} (поклади у папку програми (поруч з .exe))"
    products = []
    seen = set()
    try:
        with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=";")
            need = ["Артикул", "Назва", "Ціна ОПТ (USD)", "Ціна Роздрібна (USD)"]
            for col in need:
                if col not in (reader.fieldnames or []):
                    return [], f"У CSV нема колонки: {col}"
            for row in reader:
                art = (row.get("Артикул") or "").strip()
                name = (row.get("Назва") or "").strip()
                opt_usd = to_float(row.get("Ціна ОПТ (USD)"))
                retail_usd = to_float(row.get("Ціна Роздрібна (USD)"))
                # URL/посилання (необов'язково)
                url = ""
                for k in ("URL", "Url", "url", "Link", "LINK", "Посилання", "посилання", "Ссылка", "ссылка"):
                    if k in (reader.fieldnames or []) and (row.get(k) or "").strip():
                        url = (row.get(k) or "").strip()
                        break
                if not art:
                    continue
                if art in seen:
                    continue
                if opt_usd <= 0:
                    continue
                if retail_usd <= 0:
                    continue
                seen.add(art)
                products.append({
                    "article": art,
                    "name": name,
                    "opt_usd": opt_usd,
                    "retail_usd": retail_usd,
                                    "url": url,
})
        if not products:
            return [], "0 товарів після фільтра. Перевір ціни (ОПТ/Роздрібна)."
        return products, ""
    except Exception as e:
        return [], f"Помилка читання CSV: {e}"
# ---------- UI ----------
from kivy.uix.behaviors import ButtonBehavior

# --- paths (works both in .py and PyInstaller .exe) ---
def app_dir() -> str:
    """Directory where user-editable files live.
    - Android: App.user_data_dir (sandboxed writable dir)
    - Desktop script: folder with this .py
    - Frozen EXE: folder with the .exe
    """
    if _IS_ANDROID:
        try:
            from kivy.app import App as _KivyApp
            app = _KivyApp.get_running_app()
            if app and getattr(app, "user_data_dir", None):
                return app.user_data_dir
        except Exception:
            pass
        return os.getcwd()
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def data_path(*parts: str) -> str:
    """Path to user-editable data file/folder shipped рядом с программой."""
    return os.path.join(app_dir(), *parts)

def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def ensure_runtime_files() -> None:
    """On Android, copy bundled data files into user_data_dir on first launch.
    On desktop, does nothing.
    """
    if not _IS_ANDROID:
        return
    try:
        from kivy.resources import resource_find
    except Exception:
        resource_find = None

    base = app_dir()
    ensure_dir(base)

    # Files that should be writable by the user/app (copied once if missing)
    required_files = [
        "export.csv",
        "shops.json",
        "settings.json",
        "cart_saved.json",
        "adapter.xlsx",
    ]

    for name in required_files:
        dst = os.path.join(base, name)
        if os.path.exists(dst):
            continue
        src = None
        if resource_find:
            src = resource_find(name)
        # If resource_find fails, try relative to current working dir
        if not src:
            cand = os.path.join(os.getcwd(), name)
            if os.path.exists(cand):
                src = cand
        if src and os.path.exists(src):
            try:
                shutil.copyfile(src, dst)
            except Exception:
                pass

    # Ensure invoices folder exists
    ensure_dir(os.path.join(base, "рахунки"))


def open_url(url: str) -> None:
    """Open URL in browser (desktop) or via Android Intent."""
    if not url:
        return
    if _IS_ANDROID:
        try:
            from jnius import autoclass
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            Intent = autoclass('android.content.Intent')
            Uri = autoclass('android.net.Uri')
            intent = Intent(Intent.ACTION_VIEW, Uri.parse(url))
            PythonActivity.mActivity.startActivity(intent)
            return
        except Exception:
            # fall back to webbrowser below
            pass
    try:
        import webbrowser as _webbrowser
        _webbrowser.open(url)
    except Exception:
        pass

class ArticleLink(ButtonBehavior, Label):
    """Clickable article label that consumes touch so the parent row won't also add the item."""
    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos):
            touch.grab(self)
            return True
        return super().on_touch_down(touch)

    def on_touch_up(self, touch):
        if touch.grab_current is self:
            touch.ungrab(self)
            if self.collide_point(*touch.pos):
                self.dispatch('on_release')
            return True
        return super().on_touch_up(touch)

KV = r"""
<HeaderCell@Label>:
    size_hint_y: None
    height: dp(28)
    text_size: self.size
    halign: "left"
    valign: "middle"
    bold: True
    color: 0, 0, 0, 1
<Cell@Label>:
    text_size: self.size
    halign: "left"
    valign: "middle"
    color: 0, 0, 0, 1
<ClickLabel@ButtonBehavior+Label>:
    color: 0,0,0,1
    halign: "center"
    valign: "middle"
    text_size: self.size

<ArticleLink>:
    color: 0, 0, 0, 1
    halign: "left"
    valign: "middle"
    text_size: self.size


<TopBtn@Button>:
    size_hint: None, None
    size: dp(36), dp(36)
    background_normal: ""
    background_down: ""
    background_color: (0.1529, 0.6784, 0.9608, 1) if self.state == 'down' else (0.3, 0.3, 0.3, 1)
    color: 1, 1, 1, 1
<BarBtn@Button>:
    size_hint_y: None
    height: dp(48)
    background_normal: ""
    background_down: ""
    background_color: (0.1529, 0.6784, 0.9608, 1) if self.state == 'down' else (0.3, 0.3, 0.3, 1)
    color: 1, 1, 1, 1
    font_size: dp(18)
# --------- PRODUCTS (верхняя таблица) ----------
<ProductRow>:
    orientation: "horizontal"
    size_hint_y: None
    height: dp(54)
    padding: dp(8), dp(6)
    spacing: dp(6)
    canvas.before:
        Color:
            rgba: (0.17, 0.17, 0.17, 1) if self.is_header else ((0.905, 0.953, 0.996, 1) if self.index % 2 == 0 else (0.733, 0.867, 0.988, 1))
        Rectangle:
            pos: self.pos
            size: self.size
    ArticleLink:
        id: art_link
        text: root.article
        size_hint_x: 0.147
        shorten: True
        shorten_from: "right"
        halign: "left"
        on_release: app.root.open_article_link(root.article)
    Cell:
        text: root.name
        size_hint_x: 0.52
        halign: "left"
        valign: "middle"
        text_size: self.width, None
    Cell:
        text: root.price_uah_text
        size_hint_x: 0.11
        halign: "right"
    Cell:
        text: root.price_retail_text
        size_hint_x: 0.11
        halign: "right"
    Cell:
        text: root.margin_text
        size_hint_x: 0.10
        halign: "right"
<ProductsRV>:
    viewclass: "ProductRow"
    bar_width: dp(12)
    scroll_type: ["bars", "content"]
    scroll_wheel_distance: dp(140)
    RecycleBoxLayout:
        default_size: None, dp(54)
        default_size_hint: 1, None
        size_hint_y: None
        height: self.minimum_height
        orientation: "vertical"
# --------- CART (нижняя таблица) ----------
<CartRow>:
    orientation: "horizontal"
    size_hint_y: None
    height: dp(54)
    padding: dp(8), dp(6)   # левее
    spacing: dp(6)          # компактнее
    canvas.before:
        Color:
            rgba: (0.17, 0.17, 0.17, 1) if self.is_header else ((0.93, 0.93, 0.93, 1) if self.index % 2 == 0 else (0.87, 0.87, 0.87, 1))
        Rectangle:
            pos: self.pos
            size: self.size
    # ВАЖНО: ширины колонок совпадают с шапкой корзины
    Cell:
        text: root.num_text
        size_hint_x: None
        width: dp(28)
        halign: "center"
    ArticleLink:
        id: art_link
        text: root.article
        size_hint_x: 0.16
        shorten: True
        shorten_from: "right"
        halign: "left"
        on_release: app.root.open_article_link(root.article)
    Cell:
        text: root.name
        size_hint_x: 0.45
        halign: "left"
        valign: "middle"
        text_size: self.width, None
    RelativeLayout:
        size_hint_x: 0.06
        ClickLabel:
            size_hint: 1, 1
            padding_x: dp(6)
            pos_hint: {'x': 0, 'y': 0}
            text: "   " + root.qty_text
            color: 0, 0, 0, 1
            halign: 'center'
            valign: 'middle'
            text_size: self.size
            on_release: app.root.begin_inline_edit(root.article, 'qty', self)
    RelativeLayout:
        size_hint_x: 0.06
        ClickLabel:
            size_hint: 1, 1
            padding_x: dp(6)
            pos_hint: {'x': 0, 'y': 0}
            text: "   " + root.pct_text
            color: 0, 0, 0, 1
            halign: 'center'
            valign: 'middle'
            text_size: self.size
            on_release: app.root.begin_inline_edit(root.article, 'pct', self)
    Cell:
        text: root.price_uah_text
        size_hint_x: 0.08
        halign: "right"
    Cell:
        text: root.price_retail_text
        size_hint_x: 0.08
        halign: "right"
    Cell:
        text: root.margin_text
        size_hint_x: 0.07
        halign: "right"
    Cell:
        text: root.total_text
        size_hint_x: 0.07
        halign: "right"
<CartRV>:
    viewclass: "CartRow"
    bar_width: dp(12)
    scroll_type: ["bars", "content"]
    scroll_wheel_distance: dp(140)
    RecycleBoxLayout:
        default_size: None, dp(54)
        default_size_hint: 1, None
        size_hint_y: None
        height: self.minimum_height
        orientation: "vertical"
<Root>:
    orientation: "vertical"
    padding: [dp(8), dp(5), dp(8), dp(8)]
    spacing: dp(6)
    canvas.before:
        Color:
            rgba: 0.807, 0.835, 0.859, 1   # #ced5db
        Rectangle:
            pos: self.pos
            size: self.size
    # --- верхняя панель (НЕ трогаем визуально) ---
    BoxLayout:
        size_hint_y: None
        height: dp(44)
        spacing: dp(6)
        BoxLayout:
            size_hint_x: None
            width: dp(544)  # магазин (-20%)  # магазин x2
            spacing: dp(6)
            TextInput:
                id: shop
                halign: "left"
                padding_y: (self.height - self.line_height) / 2
                padding_x: dp(10)
                hint_text: "Магазин"
                multiline: False
                size_hint_y: None
                height: dp(36)
                font_size: dp(18)
                background_normal: ""
                background_active: ""
                background_color: 1, 1, 1, 1
                on_focus: root._select_all_on_focus(self, self.focus); root.on_shop_focus(self.focus)
                on_text: root.on_shop_text(self.text)
                on_touch_up: root._ti_touch_up_keep_selection(self, args[1])
            TopBtn:
                text: "+"
                size_hint_x: None
                width: dp(36)
                size_hint_y: None
                height: dp(36)
                font_size: dp(22)
                on_release: root.open_add_shop_popup()
            TopBtn:
                text: "..."
                size_hint_x: None
                width: dp(36)
                size_hint_y: None
                height: dp(36)
                font_size: dp(20)
                on_release: root.open_manage_shops_popup()
        TextInput:
            size_hint_y: None
            height: dp(36)
            id: search
            halign: "left"
            padding_y: (self.height - self.line_height) / 2
            padding_x: dp(10)
            hint_text: "Пошук (артикул/назва)"
            multiline: False
            size_hint_x: None
            width: dp(294)
            font_size: dp(18)
            background_normal: ""
            background_active: ""
            background_color: 1, 1, 1, 1
            on_text: root.on_search(self.text)
        TextInput:
            size_hint_y: None
            height: dp(36)
            id: kurs
            halign: "center"
            padding_y: (self.height - self.line_height) / 2
            padding_x: dp(10)
            hint_text: "Курс $"
            multiline: False
            input_filter: "float"
            size_hint_x: None
            width: dp(75)
            font_size: dp(18)
            background_normal: ""
            background_active: ""
            background_color: 1, 1, 1, 1
            halign: "center"
            on_text: root.refresh_all()
        TextInput:
            size_hint_y: None
            height: dp(36)
            id: pct
            on_focus: root._select_all_on_focus(self, self.focus)
            on_touch_up: root._pct_touch_up(self, args[1])
            padding_y: (self.height - self.line_height) / 2
            padding_x: dp(10)
            hint_text: "%"
            multiline: False
            input_filter: "float"
            size_hint_x: None
            width: dp(52)
            font_size: dp(18)
            background_normal: ""
            background_active: ""
            background_color: 0.95, 0.95, 0.95, 1
            padding: [dp(10), dp(8), dp(10), dp(8)]
            halign: "center"
            on_text: root.refresh_all()
        Label:
            id: status_inline
            text: root.status_text
            size_hint_x: 1
            size_hint_y: None
            height: dp(36)
            halign: "left"
            valign: "bottom"
            text_size: self.size
            color: 0, 0, 0, 1
        Widget:
            size_hint_x: None
            width: dp(6)

        TopBtn:
            text: root.icon_save
            font_name: root.icon_font
            font_size: dp(20)
            on_release: root.save_cart()
        TopBtn:
            text: root.icon_open
            font_name: root.icon_font
            font_size: dp(20)
            on_release: root.open_cart()
    Widget:
        size_hint_y: None
        height: dp(4)
    # шапка продуктов
    RelativeLayout:
        size_hint_y: None
        height: dp(30)
        BoxLayout:
            size_hint: None, None
            pos: -root.padding[0], 0
            size: (self.parent.width + root.padding[0] + root.padding[2]), self.parent.height
            padding: dp(8), 0
            spacing: dp(6)
            HeaderCell:
                text: "Артикул"
                size_hint_x: 0.147
            HeaderCell:
                text: "Назва"
                size_hint_x: 0.52
            HeaderCell:
                text: "Ціна за 1 шт."
                size_hint_x: 0.11
                halign: "right"
            HeaderCell:
                text: "Ціна роз. за 1 шт."
                size_hint_x: 0.11
                halign: "right"
            HeaderCell:
                text: "Маржа за 1шт."
                size_hint_x: 0.10
                halign: "right"
    # таблица товаров
    # таблица товаров (тянем зебру до краёв, без вылета вправо)
    RelativeLayout:
        size_hint_y: None
        height: root.products_h
        ProductsRV:
            id: products_rv
            size_hint: None, None
            # RelativeLayout уже внутри padding Root (имеет width = W - padL - padR).
            # Поэтому расширяем RV до полной ширины окна: parent.width + padL + padR
            pos: -root.padding[0], 0
            size: (self.parent.width + root.padding[0] + root.padding[2]), self.parent.height
    # ---------------- КОРЗИНА (меняем только её) ----------------
    CartHeader:
        size_hint_y: None
        height: dp(48)
        padding: dp(0), 0   # левее
        spacing: dp(6)      # компактнее
        HeaderCell:
            text: "№"
            size_hint_x: None
            width: dp(28)
            halign: "center"
        HeaderCell:
            text: "Артикул"
            size_hint_x: 0.16
        HeaderCell:
            text: "Назва"
            size_hint_x: 0.45
        HeaderCell:
            text: "   К-сть"
            size_hint_x: 0.06
            halign: "center"
        HeaderCell:
            text: "   %"
            size_hint_x: 0.06
            halign: "center"
        HeaderCell:
            text: "Ціна"
            size_hint_x: 0.08
            halign: "right"
        HeaderCell:
            text: "Роз."
            size_hint_x: 0.08
            halign: "right"
        HeaderCell:
            text: "Маржа"
            size_hint_x: 0.07
            halign: "right"
        HeaderCell:
            text: "Сума"
            size_hint_x: 0.07
            halign: "right"
    # таблица корзины (тянем зебру до краёв, без вылета вправо)
    RelativeLayout:
        size_hint_y: None
        height: root.cart_h
        CartRV:
            id: cart_rv
            size_hint: None, None
            pos: -root.padding[0], 0
            size: (self.parent.width + root.padding[0] + root.padding[2]), self.parent.height
    # --- нижняя панель счёта (кнопка + коментарі + итоги) ---
    BoxLayout:
        size_hint_y: None
        height: dp(36)
        padding: dp(0), dp(2)
        spacing: dp(6)
        BarBtn:
            text: root.icon_new_doc
            font_name: root.icon_font_new
            font_size: dp(20)
            size_hint: None, None
            size: dp(36), dp(36)
            pos_hint: {"center_y": .5}
            on_release: root.new_account()

        # Поле коментарів (між кнопкою очищення і підсумками)
        TextInput:
            id: order_comment
            hint_text: "Коментарі до замовлення:"
            multiline: False
            size_hint_x: 1
            size_hint_y: None
            height: dp(36)
            pos_hint: {"center_y": .5}
            font_size: dp(18)
            halign: "left"
            padding_y: (self.height - self.line_height) / 2
            padding_x: dp(10)
            background_normal: ""
            background_active: ""
            background_color: 1, 1, 1, 1
            on_focus: root._comment_clear_selection_on_focus(self, self.focus)
            on_touch_up: root._comment_clear_selection_touch(self, args[1])

        Label:
            text: "К-сть товарів:"
            size_hint_x: None
            width: dp(120)
            halign: "right"
            valign: "middle"
            text_size: self.size
            color: 0,0,0,1
        Label:
            text: root.cart_total_qty_text
            size_hint_x: None
            width: dp(60)
            halign: "left"
            valign: "middle"
            text_size: self.size
            color: 0,0,0,1
        Label:
            text: "Сума:"
            size_hint_x: None
            width: dp(55)
            halign: "right"
            valign: "middle"
            text_size: self.size
            color: 0,0,0,1
        Label:
            text: root.cart_total_sum_text
            size_hint_x: None
            width: dp(110)
            halign: "left"
            valign: "middle"
            text_size: self.size
            color: 0,0,0,1
"""
class InvoiceRow(RecycleDataViewBehavior, BoxLayout):
    path = StringProperty("")
    shop = StringProperty("")
    fname = StringProperty("")
    mtime_text = StringProperty("")
    index = NumericProperty(0)
    selected = BooleanProperty(False)

    def refresh_view_attrs(self, rv, index, data):
        self.index = index
        self.path = data.get("path", "")
        self.shop = data.get("shop", "")
        self.fname = data.get("fname", "")
        self.mtime_text = data.get("mtime_text", "")
        self.selected = bool(data.get("selected", False))
        return super().refresh_view_attrs(rv, index, data)

    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos):
            rv = self.parent.parent if hasattr(self.parent, "parent") else None
            try:
                rv = self.parent.parent
            except Exception:
                rv = None
            if rv and hasattr(rv, "select_index"):
                rv.select_index(self.index)
            return True
        return super().on_touch_down(touch)


class InvoicesRV(RecycleView):
    # holds full (unfiltered) list in ._all
    selected_path = StringProperty("")
    _all = ListProperty([])
    _selected_index = NumericProperty(-1)

    def set_all(self, rows):
        self._all = rows or []
        self.apply_filter(query="", date_from="", date_to="")

    def select_index(self, idx: int):
        try:
            idx = int(idx)
        except Exception:
            return
        self._selected_index = idx
        # mark selection in current data list
        new_data = []
        for i, d in enumerate(self.data):
            nd = dict(d)
            nd["selected"] = (i == idx)
            new_data.append(nd)
        self.data = new_data
        if 0 <= idx < len(self.data):
            self.selected_path = self.data[idx].get("path", "")

    def apply_filter(self, query: str, date_from: str, date_to: str):
        q = (query or "").strip().lower()

        def parse_date(s):
            s = (s or "").strip()
            if not s:
                return None
            try:
                # YYYY-MM-DD
                return datetime.datetime.strptime(s, "%Y-%m-%d").date()
            except Exception:
                return None

        d_from = parse_date(date_from)
        d_to = parse_date(date_to)

        out = []
        for item in (self._all or []):
            shop = item.get("shop", "")
            fname = itemz = item.get("fname", "")
            path = item.get("path", "")
            mdate = item.get("mdate")  # date object
            if q:
                hay = f"{shop} {fname}".lower()
                if q not in hay:
                    continue
            if d_from and (mdate is None or mdate < d_from):
                continue
            if d_to and (mdate is None or mdate > d_to):
                continue
            out.append(item)

        # zebra + reset selection
        self._selected_index = -1
        self.selected_path = ""
        self.data = [dict(x, selected=False) for x in out]


# KV rules for invoice list rows (lightweight, fast)
Builder.load_string(r'''
<InvoiceRow>:
    orientation: "horizontal"
    size_hint_y: None
    height: dp(54)
    padding: dp(10), dp(6)
    spacing: dp(10)
    canvas.before:
        Color:
            rgba: (0.1529, 0.6784, 0.9608, 0.75) if self.selected else ((0.27,0.27,0.27,1) if self.index % 2 == 0 else (0.35,0.35,0.35,1))
        Rectangle:
            pos: self.pos
            size: self.size
    Label:
        text: root.mtime_text
        color: 1,1,1,1
        size_hint_x: 0.22
        halign: "left"
        valign: "middle"
        text_size: self.size
    Label:
        text: root.shop
        color: 1,1,1,1
        size_hint_x: 0.30
        halign: "left"
        valign: "middle"
        text_size: self.size
    Label:
        text: root.fname
        color: 1,1,1,1
        size_hint_x: 0.48
        halign: "left"
        valign: "middle"
        shorten: True
        shorten_from: "right"
        text_size: self.size
<InvoicesRV>:
    viewclass: "InvoiceRow"
    bar_width: dp(10)
    scroll_type: ["bars", "content"]
    scroll_wheel_distance: dp(140)
    RecycleBoxLayout:
        default_size: None, dp(54)
        default_size_hint: 1, None
        size_hint_y: None
        height: self.minimum_height
        orientation: "vertical"
''')

class CartHeader(BoxLayout):
    """Шапка корзины: тянем вверх/вниз, чтобы менять высоту каталога/корзины.
    Важно: шапка не должна триггерить добавление товара в каталоге при отпускании мыши.
    """
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # icon font availability
        if _MDL2_OK:
            self.icon_font = "mdl2"
            self.icon_save = "\uE74E"   # Save
            self.icon_open = "\uE8E5"   # OpenFile
        # bottom-bar new document icon
        if _MDL2_OK:
            self.icon_new_doc = "\uE7C3"
            self.icon_font_new = "mdl2"
        else:
            self.icon_new_doc = "+"
            self.icon_font_new = "Roboto"
        self._drag_uid = None
        self._drag_start_y = 0
        self._start_cart_h = 0
        self._total_h = 0
        self._dragging = False
    def _root(self):
        p = self.parent
        while p is not None and not isinstance(p, Root):
            p = p.parent
        return p
    def on_touch_down(self, touch):
        # реагируем только если тапнули именно по шапке
        if not self.collide_point(*touch.pos):
            return super().on_touch_down(touch)
        root = self._root()
        if not root:
            return True
        self._drag_uid = touch.uid
        self._drag_start_y = touch.y
        self._start_cart_h = root.cart_h
        self._total_h = root.products_h + root.cart_h
        self._dragging = False
        # чтобы release не улетел в каталог (и не добавил товар)
        touch.grab(self)
        touch.ud["cart_header"] = True
        return True
    def on_touch_move(self, touch):
        if touch.grab_current is not self or touch.uid != self._drag_uid:
            return super().on_touch_move(touch)
        root = self._root()
        if not root:
            return True
        dy = touch.y - self._drag_start_y
        # начинаем реальный drag только после небольшого порога (чтобы тап не считался перетаскиванием)
        if not self._dragging and abs(dy) < dp(6):
            return True
        self._dragging = True
        # В Kivy y растёт вверх. Хотим: тянем вверх -> корзина больше.
        desired_cart_h = self._start_cart_h + dy
        # Во время drag делаем только лёгкие операции (без refresh/layout),
        # но важно удерживать короткую корзину "приклеенной" к шапке,
        # чтобы строки тянулись вместе с шапкой и не было скачка при отпускании.
        root.apply_split_from_drag(desired_cart_h, total_h=self._total_h, dragging=True)
        return True
    def on_touch_up(self, touch):
        # Если жест начался не на этом виджете — не трогаем
        if touch.grab_current is not self and not self.collide_point(*touch.pos):
            return super().on_touch_up(touch)
        root = self._root()
        # Отпускаем захват
        if touch.grab_current is self:
            touch.ungrab(self)
        # Если был реальный drag — фиксируем раскладку один раз и выходим
        if self._dragging:
            if root:
                root.finish_split_drag()
            return True
        # Иначе это просто тап по шапке: ничего не делаем
        return True
    # safety: legacy hook (should not be called; kept to avoid crashes if referenced in KV)
    def _load_shops(self, *args, **kwargs):
        return
class ProductsRV(RecycleView):
    pass
class CartRV(RecycleView):
    pass
class ProductRow(RecycleDataViewBehavior, BoxLayout):
    is_header = BooleanProperty(False)
    index = NumericProperty(0)
    article = StringProperty("")
    name = StringProperty("")
    price_uah_text = StringProperty("")
    price_retail_text = StringProperty("")
    margin_text = StringProperty("")
    root_ref = ObjectProperty(None)  # ссылка на Root
    def refresh_view_attrs(self, rv, index, data):
        self.index = index
        self.article = data["article"]
        self.name = data["name"]
        self.price_uah_text = data["price_uah_text"]
        self.price_retail_text = data["price_retail_text"]
        self.margin_text = data["margin_text"]
        self.root_ref = data.get("root_ref")
        return super().refresh_view_attrs(rv, index, data)

    def on_touch_up(self, touch):
        # Обрабатываем только тапы внутри строки
        if not self.collide_point(*touch.pos):
            return super().on_touch_up(touch)

        # Если это был скролл/драг — ничего не делаем
        dx = abs(touch.x - touch.opos[0])
        dy = abs(touch.y - touch.opos[1])
        if dx > dp(10) or dy > dp(10):
            return super().on_touch_up(touch)

        # СНАЧАЛА отдаём событие детям (например, ArticleLink).
        # Если ребёнок обработал — строка НЕ добавляет товар в счёт.
        if super().on_touch_up(touch):
            return True

        # Если тап именно по колонке Артикул (ArticleLink), то НЕ добавляем в счёт —
        # только открываем ссылку в браузере.
        try:
            artw = self.ids.get("art_link") if hasattr(self, "ids") else None
            if artw and artw.collide_point(*touch.pos):
                root = App.get_running_app().root
                article = getattr(self, "article", "")
                if article:
                    Clock.schedule_once(lambda dt: root.open_article_link(article), 0)
                return True
        except Exception:
            # не валим приложение из-за клика
            return True

        # Заголовок таблицы не кликается
        if getattr(self, "is_header", False):
            return True

        # Тап по строке товара = добавить в корзину
        try:
            root = App.get_running_app().root
            article = getattr(self, "article", "")
            if article:
                root._close_inline_editor(commit=True)
                Clock.schedule_once(lambda dt: root.add_to_cart(article), 0)
            return True
        except Exception as e:
            Logger.exception("ProductRow click handler failed: %s", e)
            return True

class CartRow(RecycleDataViewBehavior, BoxLayout):
    is_header = BooleanProperty(False)
    index = NumericProperty(0)
    num_text = StringProperty("")
    article = StringProperty("")
    name = StringProperty("")
    qty_text = StringProperty("")
    pct_text = StringProperty("")
    price_uah_text = StringProperty("")
    price_retail_text = StringProperty("")
    margin_text = StringProperty("")
    sum_text = StringProperty("")
    editing = BooleanProperty(False)
    field = StringProperty('')
    total_text = StringProperty('')
    # Иногда в данных строки используется поле total_text (итог/сумма).
    # RecycleView устанавливает значения по ключам data, поэтому свойство
    # должно существовать в viewclass, иначе будет AttributeError.
    total_text = StringProperty("")
    bg = ListProperty([1, 1, 1, 1])
    # ссылка на Root (для kv/inline)
    root_ref = ObjectProperty(None)
    def refresh_view_attrs(self, rv, index, data):
        self.index = index
        return super().refresh_view_attrs(rv, index, data)
class Root(BoxLayout):
    # icon glyphs / fonts for KV (must be Properties so KV can access)
    icon_new_doc = StringProperty("")
    icon_font_new = StringProperty("")
    def _icon_edit(self):
        # Pencil / edit icon
        if _MDL2_OK:
            return "\ue70f"  # Edit (pencil) in Segoe MDL2 Assets
        return "Ред"
    def _icon_del(self):
        # Trash / delete icon
        if _MDL2_OK:
            return "\ue74d"  # Delete
        return "Del"
    def _select_all_on_focus(self, ti, focused):
        """Select all text when a TextInput receives focus (for quick overwrite)."""
        if focused:
            Clock.schedule_once(lambda dt: self._do_select_all(ti), 0)
    def _ti_touch_up_keep_selection(self, ti, touch):
        # Keep "select all" highlight after click (Kivy may clear selection on touch up)
        try:
            if ti.focus and ti.collide_point(*touch.pos):
                Clock.schedule_once(lambda dt: self._select_all_on_focus(ti, True), 0)
        except Exception as e:
            self._notify('Помилка', str(e))
    def _pct_touch_up(self, ti, touch):
        # Keep full selection after mouse/touch release (TextInput normally sets cursor on touch_up)
        if not touch:
            return
        if ti.collide_point(*touch.pos):
            def _do(_dt):
                if getattr(ti, "focus", False):
                    try:
                        ti.cursor = (0, 0)
                        ti.select_all()
                    except Exception:
                        pass
            Clock.schedule_once(_do, 0)

    # -------- comment field: prevent "select all" WITHOUT moving cursor --------
    def _comment_clear_selection(self, ti):
        """If Kivy selected all text, clear selection but keep cursor."""
        try:
            if ti is None:
                return
            if getattr(ti, 'selection_text', ''):
                try:
                    ti.cancel_selection()
                except Exception:
                    try:
                        ci = ti.cursor_index()
                        ti.selection_from = ci
                        ti.selection_to = ci
                    except Exception:
                        pass
        except Exception:
            pass

    def _comment_clear_selection_on_focus(self, ti, focused):
        if focused:
            Clock.schedule_once(lambda dt: self._comment_clear_selection(ti), 0)

    def _comment_clear_selection_touch(self, ti, touch):
        try:
            if touch and ti and ti.collide_point(*touch.pos):
                Clock.schedule_once(lambda dt: self._comment_clear_selection(ti), 0)
        except Exception:
            pass

    def _do_select_all(self, ti):
        try:
            ti.select_all()
            ti.cursor = (0, 0)
        except Exception as e:
            self._notify('Помилка', str(e))
    CART_DEFAULT_PCT = 15.0  # cart rows default percent; top % affects catalog only
    status_text = StringProperty("")
    cart_total_qty_text = StringProperty("0")
    cart_total_sum_text = StringProperty("0.00")
    # --- icon-font buttons (Save/Open/Print) ---
    icon_font = StringProperty("")   # "mdl2" when available, else default font
    icon_save = StringProperty("S")
    icon_open = StringProperty("O")
    products_h = NumericProperty(0)
    cart_h = NumericProperty(0)
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # icon font availability
        if _MDL2_OK:
            self.icon_font = "mdl2"
            self.icon_save = "\uE74E"   # Save
            self.icon_open = "\uE8E5"   # OpenFile
            self.icon_new_doc = "\uE7C3"  # New document
            self.icon_font_new = "mdl2"
        else:
            # Fallback if MDL2 font not available
            self.icon_font = "Roboto"
            self.icon_save = "S"
            self.icon_open = "O"
            self.icon_new_doc = "+"
            self.icon_font_new = "Roboto"
        self.products_all = []
        self.products_filtered = []
        self.cart = {}  # article -> {"qty": int, "product": product_dict}
        self._cart_dirty = False  # True if cart changed since last save/open
        # --- shops (магазины) ---
        self.shops = []  # list of dicts: {name, phone?, note?}
        self._shop_dd = None
        self._pending_shop_text = ""
        self.selected_shop = ""
        Clock.schedule_once(lambda dt: self._load_shops_and_restore(), 0)
        Clock.schedule_once(lambda dt: self._init_defaults(), 0)
        # Apply saved курс/% after defaults & KV have settled (prevents 'random' resets).
        Clock.schedule_once(lambda dt: self._apply_ui_state(), 0.05)
        # Служебные флаги для плавного перетаскивания разделителя
        self._drag_cart_fit = False  # во время drag: корзина помещается целиком (не скроллится)
        self._drag_cart_fit_inited = False


    # -------- UI state persistence (курс / %) --------
    def _ui_state_path(self) -> str:
        """Путь для сохранения 'курс' и '%' между запусками.
        - На Android: App.user_data_dir (безопасно и всегда доступно).
        - На Windows/Linux: ~/.kivy/planet_ui_state.json (чтобы не зависеть от папки проекта).
        """
        try:
            from kivy.utils import platform as _platform
        except Exception:
            _platform = ""

        # 1) Android: user_data_dir
        if _platform == "android":
            try:
                from kivy.app import App as _KivyApp
                app = _KivyApp.get_running_app()
                base = getattr(app, "user_data_dir", "") or ""
            except Exception:
                base = ""
            if base:
                return os.path.join(base, "ui_state.json")

        # 2) Desktop: ~/.kivy
        try:
            home = os.path.expanduser("~")
            base = os.path.join(home, ".kivy")
            os.makedirs(base, exist_ok=True)
            return os.path.join(base, "planet_ui_state.json")
        except Exception:
            # 3) Fallback: рядом с файлом
            base = app_dir()
            return os.path.join(base, "ui_state.json")


    def _load_ui_state(self) -> dict:
        try:
            p = self._ui_state_path()
            if os.path.exists(p):
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    try:
                        Logger.info(f"UIState: loaded from {p} -> {data}")
                    except Exception:
                        pass
                    return data
        except Exception as e:
            try:
                Logger.exception(f"UIState: load failed: {e}")
            except Exception:
                pass
        return {}

    def _save_ui_state(self, *_):
        try:
            data = {
                "kurs": getattr(self.ids.get("kurs", None), "text", ""),
                "pct": getattr(self.ids.get("pct", None), "text", ""),
            }
            p = self._ui_state_path()
            # гарантируем папку
            base_dir = os.path.dirname(p)
            if base_dir:
                os.makedirs(base_dir, exist_ok=True)
            with open(p, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            try:
                Logger.info(f"UIState: saved to {p} -> {data}")
            except Exception:
                pass
        except Exception as e:
            try:
                Logger.exception(f"UIState: save failed: {e}")
            except Exception:
                pass

    def _schedule_save_ui_state(self, *_):
        try:
            if not hasattr(self, "_save_ui_state_trigger"):
                self._save_ui_state_trigger = Clock.create_trigger(self._save_ui_state, 0.35)
            self._save_ui_state_trigger()
        except Exception:
            pass



    def _apply_ui_state(self, *_):
        """Apply saved UI state (kurs / pct) to the top inputs.

        Called multiple times (immediately and slightly later) to override KV/default
        initialization order safely.
        """
        try:
            st = self._load_ui_state() or {}
        except Exception:
            st = {}

        # Support both direct ids and nested (just in case)
        def _set_if_present(widget_id, key):
            try:
                if widget_id in self.ids:
                    w = self.ids[widget_id]
                    v = str(st.get(key, "")).strip()
                    if v:
                        # keep user formatting; if they stored with %, keep it; if not, keep as number
                        w.text = v
            except Exception:
                pass

        _set_if_present("kurs", "kurs")
        _set_if_present("pct", "pct")

    def _init_defaults(self):
        # Set defaults only if fields are empty; do NOT overwrite values restored from ui_state.json
        try:
            if not str(self.ids.kurs.text).strip():
                self.ids.kurs.text = "43.70"
            if not str(self.ids.pct.text).strip():
                self.ids.pct.text = "15"
        except Exception:
            pass
        self._setup_split()
        self.reload_products()
    # -------- split (каталог / корзина) --------
    def _fixed_vertical_space(self):
        # Сумма всех фиксированных высот/отступов в Root (в dp).
        # ВАЖНО: значения синхронизированы с KV, иначе split может вести себя странно.
        # KV Root.padding: [dp(8), dp(5), dp(8), dp(8)] => top=5, bottom=8
        padding_tb = dp(5) + dp(8)
        # Кол-во промежутков Root.spacing между дочерними виджетами Root.
        # Сейчас блоков 8: верх, отступ(4px), статус, шапка каталога, каталог, шапка кошика, кошик, нижняя панель
        spacing_count = 6
        spacings = dp(6) * spacing_count  # Root.spacing = dp(6)
        top_panel = dp(44)      # верхняя панель
        gap = dp(4)             # небольшой отступ (Widget height)
        # статус перенесён в верхнюю панель
        prod_header = dp(30)    # шапка товаров
        prod_header = dp(30)    # шапка товаров
        cart_header = dp(48)    # шапка корзины
        bottom_bar = dp(32)     # нижняя панель счёта
        return padding_tb + spacings + top_panel + gap + prod_header + cart_header + bottom_bar
    def _available_height(self):
        return max(0, Window.height - self._fixed_vertical_space())
    def _setup_split(self):
        # Стартовые пропорции такие же, как были по size_hint_y: 0.58 и 0.30
        total = self._available_height()
        if total <= 0:
            self.products_h = dp(200)
            self.cart_h = dp(150)
        else:
            w_prod, w_cart = 0.58, 0.30
            cart = total * (w_cart / (w_prod + w_cart))
            prod = total - cart
            self.products_h = prod
            self.cart_h = cart
        # clamp по минимумам (по 2 строки)
        self.apply_split_from_drag(self.cart_h, self.products_h + self.cart_h)
        Window.unbind(on_resize=self._on_window_resize)
        Window.bind(on_resize=self._on_window_resize)
    def _on_window_resize(self, window, w, h):
        # сохраняем текущую долю корзины
        total_old = self.products_h + self.cart_h
        total_new = self._available_height()
        if total_old <= 0 or total_new <= 0:
            return
        cart_ratio = self.cart_h / total_old
        self.apply_split_from_drag(total_new * cart_ratio, total_new)
    def apply_split_from_drag(self, desired_cart_h, total_h=None, dragging: bool = False):
        """
        Меняет высоты каталога/корзины при перетаскивании шапки корзины.
        Главная причина "дёрганья" у верхнего упора — когда высоты уже упёрлись в clamp,
        но мы продолжаем каждое движение мыши/пальца снова присваивать те же значения.
        Каждое присваивание инвалидирует layout и даёт микроснап.
        Поэтому:
          1) клампим + округляем к целым пикселям
          2) если итоговые значения не изменились — выходим (ничего не перерисовываем)
          3) во время drag делаем только лёгкую коррекцию scroll_y (без refresh/do_layout)
        """
        if total_h is None:
            total_h = self.products_h + self.cart_h
        total_h = max(0.0, float(total_h))
        # минимум 2 строки видно и там, и там
        row_h = dp(54)
        min_prod = row_h * 2
        min_cart = row_h * 2
        max_cart = max(min_cart, total_h - min_prod)
        new_cart = min(max(float(desired_cart_h), float(min_cart)), float(max_cart))
        new_prod = max(float(min_prod), total_h - new_cart)
        # держим total_h постоянным
        new_cart = total_h - new_prod
        # округление к "пикселю" (сильно снижает дребезг у упоров)
        new_cart = float(int(round(new_cart)))
        new_cart = min(max(new_cart, float(min_cart)), float(max_cart))
        new_prod = float(int(round(total_h - new_cart)))
        if new_prod < float(min_prod):
            new_prod = float(int(round(min_prod)))
            new_cart = float(int(round(total_h - new_prod)))
        # если фактически ничего не поменялось — не трогаем свойства (это и убирает дёрганье)
        if abs(new_cart - float(self.cart_h)) < 0.5 and abs(new_prod - float(self.products_h)) < 0.5:
            if dragging:
                # но лёгкую "прилипалку" корзины всё равно поддержим
                self._keep_cart_attached_to_header_light()
            return
        self.cart_h = new_cart
        self.products_h = total_h - new_cart
        if dragging:
            self._keep_cart_attached_to_header_light()
    def _keep_cart_attached_to_header_light(self):
        """Лёгкая коррекция во время drag.
        Если строк мало и они помещаются — держим корзину "кверху" (scroll_y = 1),
        но делаем это только при переходе в состояние "помещается", чтобы не было микроснапов.
        """
        try:
            rv = self.ids.get("cart_rv")
            if not rv:
                return
            row_h = dp(54)
            content_h = (len(getattr(rv, "data", []) or []) * row_h)
            viewport_h = float(self.cart_h)
            fits = content_h <= viewport_h + dp(1)
            # Инициализируем состояние в начале drag
            if not getattr(self, "_drag_cart_fit_inited", False):
                self._drag_cart_fit = fits
                self._drag_cart_fit_inited = True
            # Если корзина стала "вмещаться" — один раз поднимем кверху
            if fits and (not self._drag_cart_fit):
                self._drag_cart_fit = True
                if rv.scroll_y != 1:
                    rv.scroll_y = 1
            elif not fits:
                self._drag_cart_fit = False
        except Exception:
            return
    def finish_split_drag(self):
        # Вызываем после окончания перетаскивания шапки (on_touch_up).
        # Тогда пересчёт происходит один раз и интерфейс остаётся плавным.
        self._drag_cart_fit_inited = False
        Clock.schedule_once(lambda dt: self._force_split_relayout(), 0)
    def _force_split_relayout(self):
        try:
            # do_layout пересчитает позиции/клиппинг
            self.do_layout()
            def _fix_rv(rv):
                lm = getattr(rv, "layout_manager", None)
                content_h = float(getattr(lm, "height", 0) or 0)
                viewport_h = float(getattr(rv, "height", 0) or 0)
                # refresh после смены высоты
                rv.refresh_from_layout()
                if content_h <= viewport_h + dp(1):
                    rv.scroll_y = 1  # при маленьком контенте всегда "кверху"
                    rv.refresh_from_layout()
            if "products_rv" in self.ids:
                _fix_rv(self.ids.products_rv)
            if "cart_rv" in self.ids:
                _fix_rv(self.ids.cart_rv)
        except Exception as e:
            self._notify('Помилка', str(e))
    def kurs(self):
        k = to_float(self.ids.kurs.text)
        return k if k > 0 else 1.0
    def pct(self):
        return to_float(self.ids.pct.text)
    
    def open_article_link(self, article: str):
        """Відкрити посилання на товар по артикулу (якщо є в базі)."""
        art = (article or "").strip()
        if not art:
            return
        url = ""
        # спроба знайти у завантажених товарах
        for p in getattr(self, "products_all", []) or []:
            if (p.get("article") or "").strip() == art:
                url = (p.get("url") or "").strip()
                break
        if not url:
            # якщо посилання нема — відкриємо пошук (щоб не було "ніщо не сталося")
            url = f"https://www.google.com/search?q={art}"
        try:
            open_url(url)
        except Exception as e:
            try:
                self._notify("Помилка", f"Не вдалося відкрити браузер:\n{e}")
            except Exception:
                pass
    def reload_products(self):
        project_dir = app_dir()
        csv_path = data_path('export.csv')
        products, err = load_products_from_export_csv(csv_path)
        self.products_all = products
        if err:
            self.status_text = f"Помилка: {err}"
            Popup(title="Помилка",
                content=Label(text=err),
                size_hint=(None, None),
                size=(int(dp(820)), int(dp(260))),
            ).open()
            self.ids.products_rv.data = []
            return
        self.status_text = f"Товарів: {len(products)}"
        self.on_search(self.ids.search.text)
    def on_search(self, text):
        q = (text or "").strip().lower()
        if not q:
            self.products_filtered = list(self.products_all)
        else:
            self.products_filtered = [
                p for p in self.products_all
                if q in p["article"].lower() or q in p["name"].lower()
            ]
        self.refresh_products()
    def refresh_all(self):
        self.refresh_products()
        self.refresh_cart()
    def refresh_products(self):
        k = self.kurs()
        pct = self.pct()
        data = []
        for p in self.products_filtered:
            price_uah = p["opt_usd"] * k
            price_retail = (p.get("retail_usd", 0.0) or 0.0) * k
            if price_retail <= 0:  # если в файле нет розничной цены, используем цену закупки
                price_retail = price_uah
            margin = price_uah * (pct / 100.0)
            data.append({
                "article": p["article"],
                "name": p["name"],
                "price_uah_text": f"{price_uah:.2f}",
                "price_retail_text": f"{price_retail:.2f}",
                "margin_text": f"{margin:.2f}",
                "root_ref": self,
            })
        self.ids.products_rv.data = data
    def refresh_cart(self):
        k = self.kurs()
        default_pct = self.CART_DEFAULT_PCT
        rows = []
        i = 1
        qty_sum = 0
        total_sum = 0.0
        for article, item in self.cart.items():
            p = item["product"]
            qty = int(item["qty"])
            # % для строки корзины хранится в самой строке (верхний % влияет только на каталог)
            row_pct = int(item.get('pct', default_pct))
            row_pct_txt = f"{row_pct:.0f}"
            price_uah = p["opt_usd"] * k
            # Розница считаем от опта + индивидуальный %
            price_retail = price_uah * (1.0 + (row_pct / 100.0))
            margin = price_retail - price_uah
            total = price_retail * qty
            qty_sum += qty
            total_sum += total
            rows.append({
                "num_text": str(i),
                "article": article,
                "name": p["name"],
                "qty_text": str(qty),
                "pct_text": row_pct_txt,
                "price_uah_text": f"{price_uah:.2f}",
                "price_retail_text": f"{price_retail:.2f}",
                "margin_text": f"{margin:.2f}",
                "total_text": f"{total:.2f}",
                "editing": (article == self.edit_article),
                "field": self.edit_field,
                "root_ref": self,
            })
            i += 1
        # totals for footer
        self.cart_total_qty_text = str(qty_sum)
        self.cart_total_sum_text = f"{total_sum:.2f}"
        self.ids.cart_rv.data = rows
        # Если после удаления/изменения позиций контент корзины стал меньше окна,
        # принудительно держим его \"прилипшим\" к шапке (без пустоты сверху).
        Clock.schedule_once(self._ensure_cart_top_if_not_scrollable, 0)
    def _ensure_cart_top_if_not_scrollable(self, _dt=0):
        """После изменения данных корзины убирает пустоту сверху, если скролл не нужен."""
        rv = self.ids.get("cart_rv")
        if not rv:
            return
        # лёгкий refresh (не тяжёлый do_layout на каждом движении)
        try:
            rv.refresh_from_layout()
        except Exception as e:
            self._notify('Помилка', str(e))
        layout = rv.children[0] if rv.children else None
        if not layout:
            return
        try:
            content_h = float(getattr(layout, "minimum_height", 0) or 0)
            viewport_h = float(rv.height or 0)
        except Exception:
            return
        if content_h <= viewport_h + 1.0:
            try:
                rv.scroll_y = 1.0
            except Exception:
                pass
# --- Inline edit for cart (qty / %) ---
    edit_article = StringProperty("")
    edit_field = StringProperty("")  # 'qty' or 'pct'

    # --- Overlay editor for cart (qty / %) ---
    _overlay_input = ObjectProperty(None, allownone=True)
    _overlay_target_article = StringProperty("")
    _overlay_target_field = StringProperty("")  # 'qty' or 'pct'
    _overlay_session = NumericProperty(0)
    _overlay_original_text = StringProperty("")
    _overlay_active = BooleanProperty(False)

    def _ensure_overlay_input(self):
        """Create a single floating TextInput used to edit qty/% in the cart.
        This avoids RecycleView focus/commit glitches of per-row TextInputs.
        """
        if self._overlay_input is not None:
            return
        ti = TextInput(
            size_hint=(None, None),
            multiline=False,
            halign="center",
            background_normal="",
            background_active="",
            background_color=(1, 1, 1, 1),
            foreground_color=(0, 0, 0, 1),
            cursor_color=(0, 0, 0, 1),
            padding=[dp(10), dp(8), dp(10), dp(8)],
            opacity=0,
            disabled=True,
        )
        ti.bind(on_text_validate=self._on_overlay_validate)
        ti.bind(focus=self._on_overlay_focus)
        self._overlay_input = ti
        try:
            Window.add_widget(ti)
        except Exception:
            pass
        # Defocus when tapping outside (do NOT swallow touch)
        try:
            Window.bind(on_touch_down=self._on_window_touch_down_overlay)
        except Exception:
            pass

    def _overlay_apply_centering(self, session_id=None):
        """Pixel-stable centering for the floating editor."""
        if session_id is not None and session_id != self._overlay_session:
            return
        ti = self._overlay_input
        if ti is None or ti.disabled or ti.opacity <= 0:
            return
        try:
            ti.halign = "center"
        except Exception:
            pass
        try:
            lh = float(getattr(ti, "line_height", 0) or 0)
            if lh <= 0:
                lh = float(getattr(ti, "font_size", 16) or 16)
            py = max(0, int((float(ti.height) - lh) / 2))
            px = int(dp(10))
            ti.padding = [px, py, px, py]
        except Exception:
            pass

    def _overlay_select_all(self, session_id):
        if not self._overlay_active or session_id != self._overlay_session:
            return
        ti = self._overlay_input
        if ti is None:
            return
        try:
            ti.select_all()
        except Exception:
            pass
        self._overlay_apply_centering(session_id)

    def begin_inline_edit(self, article: str, field: str, anchor_widget=None):
        """Open the floating editor over the clicked cell."""
        if field not in ("qty", "pct"):
            return
        if not anchor_widget:
            return
        if article not in self.cart:
            return

        self._ensure_overlay_input()

        self.edit_article = article
        self.edit_field = field

        if field == "qty":
            cur_text = str(int(self.cart[article].get("qty", 1)))
            self._overlay_input.input_filter = "int"
        else:
            cur_text = str(int(self.cart[article].get("pct", int(self.CART_DEFAULT_PCT))))
            self._overlay_input.input_filter = "int"

        self._overlay_session += 1
        self._overlay_target_article = article
        self._overlay_target_field = field
        self._overlay_original_text = cur_text
        self._overlay_active = True

        try:
            x, y = anchor_widget.to_window(anchor_widget.x, anchor_widget.y)
            w, h = anchor_widget.size
            # Keep overlay editor aligned with visually indented label text (prevents 'jump' on tap)
            if field in ("qty", "pct"):
                indent = dp(12)
                x += indent
                w = max(w - indent, dp(54))
            w = max(w, dp(54))
            h = max(h, dp(36))
            # Center vertically within the cell to eliminate the ~15px shift
            cell_center_y = y + h / 2.0
            editor_h = max(dp(36), h)
            editor_y = cell_center_y - editor_h / 2.0
            # Clamp inside visible cart area (so editor doesn't go under footer / outside view)
            # Also: do NOT activate editor if the clicked cell is not fully visible (prevents "top partially hidden still activates")
            try:
                cart_rv = self.ids.get("cart_rv")
                if cart_rv is not None:
                    cx, cy = cart_rv.to_window(cart_rv.x, cart_rv.y)
                    ctop = cy + cart_rv.height
                    cbottom = cy
                    margin = dp(2)
                    # If the cell is clipped by the cart view (top/bottom), ignore the tap (user expects no activation)
                    if y < cbottom + margin or (y + h) > ctop - margin:
                        self._close_inline_editor(commit=False)
                        return
                    editor_y = min(editor_y, ctop - editor_h - margin)
                    editor_y = max(editor_y, cbottom + margin)
            except Exception:
                pass
            self._overlay_input.size = (w, editor_h)
            self._overlay_input.pos = (x, editor_y)
        except Exception:
            self._overlay_input.size = (dp(80), dp(36))
            self._overlay_input.pos = (dp(10), dp(10))

        ti = self._overlay_input
        ti.text = cur_text
        ti.opacity = 1
        ti.disabled = False
        ti.focus = True

        Clock.schedule_once(lambda dt: self._overlay_apply_centering(self._overlay_session), 0)
        Clock.schedule_once(lambda dt: self._overlay_apply_centering(self._overlay_session), 0.02)
        Clock.schedule_once(lambda dt: self._overlay_apply_centering(self._overlay_session), 0.06)
        Clock.schedule_once(lambda dt: self._overlay_select_all(self._overlay_session), 0)
        Clock.schedule_once(lambda dt: self._overlay_select_all(self._overlay_session), 0.02)
        Clock.schedule_once(lambda dt: self._overlay_select_all(self._overlay_session), 0.06)

    def _on_window_touch_down_overlay(self, window, touch):
        if not self._overlay_active or self._overlay_input is None:
            return False
        ti = self._overlay_input
        if ti.disabled or ti.opacity <= 0:
            return False
        if not ti.collide_point(*touch.pos):
            try:
                ti.focus = False
            except Exception:
                pass
        return False

    def _on_overlay_validate(self, ti):
        self._commit_overlay(ti.text)

    def _on_overlay_focus(self, ti, focused):
        if focused:
            return
        if self._overlay_active:
            self._commit_overlay(ti.text)

    
    # --- Compatibility: old code paths called _close_inline_editor(commit=...)
    # With overlay editor we just commit/hide the overlay if it's active.
    def _close_inline_editor(self, commit: bool = True):
        try:
            if getattr(self, "_overlay_active", False):
                ti = getattr(self, "_overlay_input", None)
                if commit and ti is not None:
                    self._commit_overlay(ti.text)
                else:
                    # just hide without committing
                    self._overlay_active = False
                    if ti is not None:
                        ti.disabled = True
                        ti.opacity = 0
                        ti.focus = False
                    self.edit_article = ""
                    self.edit_field = ""
        except Exception:
            # never crash on close helper
            pass

    def _commit_overlay(self, text_value: str):
        if not self._overlay_active:
            return
        article = self._overlay_target_article
        field = self._overlay_target_field
        ti = self._overlay_input

        self._overlay_active = False
        if ti is not None:
            ti.disabled = True
            ti.opacity = 0
            ti.focus = False

        v = (text_value or "").strip()
        orig = (self._overlay_original_text or "").strip()
        if v == "":
            v = orig

        try:
            num = int(float(v))
        except Exception:
            try:
                num = int(float(orig))
            except Exception:
                num = 0

        if article not in self.cart:
            self.edit_article = ""
            self.edit_field = ""
            return

        changed = False
        if field == "qty":
            old_num = int(self.cart[article].get("qty", 1))
            if num <= 0:
                self.cart.pop(article, None)
                changed = True
            else:
                if num != old_num:
                    self.cart[article]["qty"] = num
                    changed = True
        else:
            old_num = int(self.cart[article].get("pct", int(self.CART_DEFAULT_PCT)))
            if num < 0:
                num = 0
            if num > 999:
                num = 999
            if num != old_num:
                self.cart[article]["pct"] = num
                changed = True

        self.edit_article = ""
        self.edit_field = ""

        if changed:
            self._cart_dirty = True
            self.refresh_cart()


    def add_to_cart(self, article: str):
        p = next((x for x in self.products_all if x["article"] == article), None)
        if not p:
            return
        # Поточний % береться з верхнього поля (id: pct) на момент додавання.
        # ВАРІАНТ B: якщо товар уже є в корзині — збільшуємо кількість і ОНОВЛЮЄМО % у позиції.
        cur_pct = int(self.CART_DEFAULT_PCT)
        try:
            if 'pct' in self.ids:
                _t = str(self.ids.pct.text).strip().replace('%', '').replace(',', '.')
                if _t:
                    cur_pct = int(float(_t))
        except Exception:
            pass

        if article in self.cart:
            self.cart[article]["qty"] += 1
            # Оновлюємо % у вже існуючої позиції (щоб наступні додавання відображали новий %).
            self.cart[article]["pct"] = cur_pct
        else:
            self.cart[article] = {'qty': 1, 'pct': cur_pct, 'product': p}

        self._cart_dirty = True
        self.refresh_cart()

    def set_cart_qty(self, article: str, qty: int):
        if not article:
            return
        if article not in self.cart:
            return
        if qty <= 0:
            # удалить позицию
            try:
                del self.cart[article]
            except Exception:
                pass
            self._cart_dirty = True
        else:
            self.cart[article]["qty"] = int(qty)
            self._cart_dirty = True
        self.refresh_cart()
    def open_qty_popup(self, article: str):
        if not article or article not in self.cart:
            return
        item = self.cart[article]
        p = item["product"]
        cur_qty = int(item.get("qty", 1))
        k = self.kurs()
        default_pct = self.CART_DEFAULT_PCT  # верхний % влияет только на каталог
        row_pct = float(item.get('pct', default_pct))
        price_uah = p['opt_usd'] * k
        price_retail = price_uah * (1.0 + row_pct / 100.0)
        margin = price_retail - price_uah
        # закрыть предыдущий, если открыт
        if getattr(self, "_qty_popup", None):
            try:
                self._qty_popup.dismiss()
            except Exception:
                pass
            self._qty_popup = None
        box = BoxLayout(orientation="vertical", spacing=dp(8), padding=[dp(10), dp(10), dp(10), dp(10)])
        info = Label(
            text=f"[b]{p['name']}[/b]\nАртикул: {article}\nПоточна к-сть: {cur_qty}\nЦіна: {price_uah:.2f} | Роз.: {price_retail:.2f} | Маржа: {margin:.2f}",
            markup=True,
            halign="left",
            valign="middle",
            text_size=(dp(860), None),
            size_hint_y=None,
        )
        # высота под 4 строки
        info.height = dp(110)
        box.add_widget(info)
        ti = TextInput(
            text=str(cur_qty),
            multiline=False,
            input_filter="int",
            font_size=dp(22),
            size_hint_y=None,
            height=dp(46),
            halign="center",
        )
        box.add_widget(ti)
        btns = GridLayout(cols=3, spacing=dp(8), size_hint_y=None, height=dp(46))
        btn_cancel = PopupButton(text="Скасувати")
        btn_delete = Button(text="Видалити")
        btn_ok = Button(text="OK")
        btns.add_widget(btn_cancel)
        btns.add_widget(btn_delete)
        btns.add_widget(btn_ok)
        box.add_widget(btns)
        popup = Popup(title="Встановити кількість",
            content=box,
            size_hint=(None, None),
            size=(int(dp(900)), int(dp(280))),
            auto_dismiss=False,
        )
        self._qty_popup = popup
        # Повторно ставим фокус/выделение (на ПК отпускание мышки может снимать выделение)
        popup.bind(on_open=lambda *_a: Clock.schedule_once(_focus, 0))
        def _focus(_dt):
            try:
                ti.focus = True
                ti.select_all()
            except Exception:
                pass
        def _close(*_a):
            try:
                popup.dismiss()
            except Exception:
                pass
            self._qty_popup = None
        def _apply(*_a):
            q = to_float(ti.text)
            # to_float вернёт float, но нам нужен int
            try:
                q_int = int(q)
            except Exception:
                q_int = cur_qty
            self.set_cart_qty(article, q_int)
            _close()
        def _del(*_a):
            self.set_cart_qty(article, 0)
            _close()
        btn_cancel.bind(on_release=_close)
        btn_ok.bind(on_release=_apply)
        btn_delete.bind(on_release=_del)
        popup.open()
        Clock.schedule_once(_focus, 0)
        Clock.schedule_once(_focus, 0.08)
        Clock.schedule_once(_focus, 0.16)
    # -------- shops (магазины) --------
    def _project_dir(self):
        return app_dir()
    def _shops_path(self):
        return os.path.join(self._project_dir(), "shops.json")
    def _settings_path(self):
        return os.path.join(self._project_dir(), "settings.json")
    def load_shops(self):
        """Public alias for shops loading (kept for backward/forward compatibility)."""
        self._load_shops()
    def _load_shops_and_restore(self):
        # Загружаем список магазинов, но НЕ подставляем прошлый выбор автоматически.
        # Стартуем с пустого поля (показывается hint_text "Магазин").
        self.load_shops()
        self.selected_shop = ""
        if "shop" in self.ids:
            self.ids.shop.text = ""
            # Оставляем фокус/курсор как есть — пользователь сам выберет магазин.

    def _load_shops(self):
        """Load shops from shops.json.

        Supported formats:
          - ["Shop A", "Shop B"]
          - [{"name": "Shop A", "phone": "...", "note": "..."}]
        """
        path = self._shops_path()
        shops = []
        try:
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)

                if isinstance(data, list):
                    for item in data:
                        if isinstance(item, str):
                            name = item.strip()
                            if name:
                                shops.append({"name": name})
                        elif isinstance(item, dict):
                            name = str(item.get("name", "")).strip()
                            if name:
                                shops.append(
                                    {
                                        "name": name,
                                        "phone": str(item.get("phone", "")).strip(),
                                        "note": str(item.get("note", "")).strip(),
                                    }
                                )
        except Exception as e:
            Logger.warning(f"shops: failed to load shops.json: {e}")

        # Sort shops alphabetically (case-insensitive) so dropdown and popup share the same order
        shops.sort(key=lambda s: (str(s.get('name','')).strip().casefold() if isinstance(s, dict) else str(s).casefold()))
        self.shops = shops
        return shops

    def _sort_shops_inplace(self):
        """Sort self.shops by name (case-insensitive)."""
        try:
            self.shops = [s for s in (self.shops or []) if isinstance(s, dict) and str(s.get('name','')).strip()]
            self.shops.sort(key=lambda s: str(s.get('name','')).strip().casefold())
        except Exception as e:
            Logger.warning(f"shops: sort failed: {e}")

    def _ensure_shop_in_list(self, shop_name: str):
        """If user typed a new shop name manually, add it to shops.json and refresh UI."""
        name = (shop_name or '').strip()
        if not name:
            return
        # Normalize existing names
        try:
            existing = {str(s.get('name','')).strip() for s in (self.shops or []) if isinstance(s, dict)}
        except Exception:
            existing = set()
        if name not in existing:
            (self.shops or []).append({'name': name, 'phone': '', 'note': ''})
            self._sort_shops_inplace()
            self._shops_file_write()
        else:
            # still keep order stable and consistent
            self._sort_shops_inplace()
        try:
            self._refresh_shops_list()
        except Exception:
            pass

    def _restore_last_shop(self):
        """Restore last selected shop from settings.json into the input field."""
        last = ""
        try:
            sp = self._settings_path()
            if os.path.exists(sp):
                with open(sp, "r", encoding="utf-8") as f:
                    data = json.load(f) or {}
                last = str(data.get("last_shop", "")).strip()
        except Exception:
            last = ""
        if last and "shop" in self.ids:
            self.selected_shop = last
            self.ids.shop.text = last
    def _save_last_shop(self, name: str):
        try:
            sp = self._settings_path()
            data = {}
            if os.path.exists(sp):
                try:
                    with open(sp, "r", encoding="utf-8") as f:
                        data = json.load(f) or {}
                except Exception:
                    data = {}
            data["last_shop"] = name
            with open(sp, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            Logger.warning(f"shops: failed to save settings.json: {e}")
    def _refresh_shops_list(self):
        """Refresh shop-related UI after add/edit/delete.
        We keep this intentionally lightweight: update the main shop input
        selection (if the selected shop was deleted) and rebuild the
        dropdown suggestions.
        """
        try:
            current = (getattr(self, "selected_shop", "") or "").strip()
            names = [
                (s.get("name", "") or "").strip()
                for s in (getattr(self, "shops", []) or [])
                if isinstance(s, dict)
            ]
            if current and current not in names:
                self.selected_shop = ""
                if hasattr(self, "ids") and "shop" in self.ids:
                    self.ids.shop.text = ""
        except Exception as e:
            Logger.warning(f"shops: refresh selection failed: {e}")
        # Rebuild dropdown based on whatever is currently typed
        try:
            if hasattr(self, "ids") and "shop" in self.ids:
                self.on_shop_text(self.ids.shop.text)
        except Exception as e:
            Logger.warning(f"shops: refresh dropdown failed: {e}")
    # ---------- Shops: add / manage UI ----------
    def _shops_file_write(self):
        """Persist current shops list to shops.json (as list of dicts)."""
        try:
            self._sort_shops_inplace()
            path = self._shops_path()
            data = []
            for s in (self.shops or []):
                if not isinstance(s, dict):
                    continue
                name = str(s.get("name", "")).strip()
                if not name:
                    continue
                data.append({
                    "name": name,
                    "phone": str(s.get("phone", "")).strip(),
                    "note": str(s.get("note", "")).strip(),
                })
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            Logger.warning(f"shops: failed to save shops.json: {e}")
    def _shop_name_exists(self, name: str, *, exclude_name: str = "") -> bool:
        n = (name or "").strip().lower()
        ex = (exclude_name or "").strip().lower()
        if not n:
            return False
        for s in (self.shops or []):
            sn = str(s.get("name", "")).strip().lower()
            if sn == n and sn != ex:
                return True
        return False
    def open_add_shop_popup(self):
        """Popup to add a new shop (prefills name from current shop input for quick add)."""
        try:
            # Take whatever user typed in the shop input (even if not found in list)
            pending = (self.ids.shop.text if "shop" in self.ids else "").strip()
        except Exception:
            pending = ""
        self._pending_shop_text = pending
        return self._open_edit_shop_popup(shop=None, is_new=True)

    def apply_shop_selection(self, shop_name: str):
        """Apply chosen shop to main input and highlight it for quick edit."""
        if not shop_name:
            return
        self.selected_shop = shop_name
        if 'shop' in self.ids:
            self.ids.shop.text = shop_name
            self.ids.shop.focus = True
            from kivy.clock import Clock
            Clock.schedule_once(lambda dt: self.ids.shop.select_all(), 0)
        try:
            self.refresh_shops_selection()
        except Exception:
            pass

    def open_manage_shops_popup(self):
        """
        Popup: список магазинов с поиском + действия (редактировать/удалить).
        Требования:
        - без кнопки "Закрити" (закрытие по клику вне окна);
        - поле поиска на всю ширину;
        - две "слепленные" квадратные кнопки справа (edit/delete) без зазоров/разделителей;
        - иконки как в главном окне (Segoe MDL2 Assets, если доступно).
        """
        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.scrollview import ScrollView
        from kivy.uix.textinput import TextInput
        from kivy.uix.button import Button
        from kivy.uix.label import Label
        from kivy.metrics import dp, sp
        from kivy.clock import Clock
        from kivy.graphics import Color, Rectangle
        from kivy.graphics.instructions import InstructionGroup
        from kivy.effects.scroll import ScrollEffect
        # Подстрахуемся: загрузим магазины, если список пуст
        try:
            if not getattr(self, "shops", None):
                self.load_shops()
        except Exception as e:
            self._notify('Помилка', str(e))
        # --- размеры (как в основном UI) ---
        row_h = dp(54)
        ti_h = dp(36)
        btn_w = row_h  # квадратные кнопки
        btn_col_w = btn_w * 2
        # --- цвета ---
        BG_PANEL = (0.18, 0.18, 0.18, 1)
        Z1 = (0.38, 0.38, 0.38, 1)
        Z2 = (0.44, 0.44, 0.44, 1)
        BTN_EDIT_UP = (0.58, 0.58, 0.58, 1)
        BTN_EDIT_DOWN = (0.52, 0.52, 0.52, 1)
        BTN_DEL_UP = (0.98, 0.52, 0.13, 1)   # оранжевый
        BTN_DEL_DOWN = (0.90, 0.45, 0.10, 1)
        ICON_FONT = "mdl2" if _MDL2_OK else None
        def _apply_bg(w, rgba):
            w.canvas.before.clear()
            with w.canvas.before:
                Color(*rgba)
                w._bg_rect = Rectangle(pos=w.pos, size=w.size)
            def _upd(*_):
                w._bg_rect.pos = w.pos
                w._bg_rect.size = w.size
            w.bind(pos=_upd, size=_upd)

        # press-highlight helpers (do not destroy zebra background)
        HL_RGBA = (0.1529, 0.6784, 0.9608, 1)  # same blue as '+' button style
        def _press_on(row):
            if getattr(row, '_press_ig', None) is not None:
                return
            ig = InstructionGroup()
            ig.add(Color(*HL_RGBA))
            rect = Rectangle(pos=row.pos, size=row.size)
            ig.add(rect)
            row._press_rect = rect
            row._press_ig = ig
            row.canvas.before.add(ig)
            def _upd_press(*_):
                if getattr(row, '_press_rect', None) is not None:
                    row._press_rect.pos = row.pos
                    row._press_rect.size = row.size
            row.bind(pos=_upd_press, size=_upd_press)

        def _press_off(row):
            ig = getattr(row, '_press_ig', None)
            if ig is None:
                return
            try:
                row.canvas.before.remove(ig)
            except Exception:
                pass
            row._press_ig = None
            row._press_rect = None
        def _style_icon_btn(btn: Button, up_rgba, down_rgba):
            btn.background_normal = ""
            btn.background_down = ""
            btn.border = (0, 0, 0, 0)
            btn.background_color = up_rgba
            btn.color = (1, 1, 1, 1)
            btn.font_size = dp(24)
            if ICON_FONT:
                btn.font_name = ICON_FONT
            def _on_state(_btn, _state):
                _btn.background_color = down_rgba if _state == "down" else up_rgba
            btn.bind(state=_on_state)
            return btn
        class _TapSelectBtn(Button):
            """Button that reliably selects on tap inside ScrollView (grabs touch)."""
            shop_name = ""
            def __init__(self, choose_cb=None, **kwargs):
                super().__init__(**kwargs)
                self._choose_cb = choose_cb
                self._down_pos = None
                self._moved = False
            def on_touch_down(self, touch):
                if self.collide_point(*touch.pos):
                    touch.grab(self)
                    self._down_pos = touch.pos
                    self._moved = False
                    # show blue press overlay
                    self.background_color = (0.1529, 0.6784, 0.9608, 1)
                    return True
                return super().on_touch_down(touch)
            def on_touch_move(self, touch):
                if touch.grab_current is self:
                    if self._down_pos:
                        dx = abs(touch.x - self._down_pos[0])
                        dy = abs(touch.y - self._down_pos[1])
                        if dx > dp(8) or dy > dp(8):
                            self._moved = True
                            # stop press highlight while scrolling
                            self.background_color = (0, 0, 0, 0)
                    return True
                return super().on_touch_move(touch)
            def on_touch_up(self, touch):
                if touch.grab_current is self:
                    touch.ungrab(self)
                    # remove highlight
                    self.background_color = (0, 0, 0, 0)
                    if (not self._moved) and self.collide_point(*touch.pos):
                        if callable(self._choose_cb):
                            self._choose_cb(self.shop_name)
                        self._close_inline_editor(commit=False)
                        return True
                    return True
                return super().on_touch_up(touch)
        # -------- Popup container (без системной рамки) --------
        outer = BoxLayout(orientation="vertical", spacing=0, padding=[0, 0, 0, 0])
        _apply_bg(outer, (0, 0, 0, 0))  # transparent outer background
        # Header: только поиск на всю ширину
        header = BoxLayout(orientation="horizontal", size_hint_y=None, height=ti_h, spacing=0, padding=[0, 0, 0, 0])
        ti = TextInput(
            hint_text="Пошук магазину",
            multiline=False,
            size_hint=(1, None),
            height=ti_h,
            padding=[dp(10), dp(8), dp(10), dp(8)],
            background_normal="",
            background_active="",
            background_color=(1, 1, 1, 1),
            foreground_color=(0, 0, 0, 1),
            cursor_color=(0, 0, 0, 1),
        )
        header.add_widget(ti)
        outer.add_widget(header)
        # Scroll area (без overscroll заливок)
        sv = ScrollView(
            do_scroll_x=False,
            bar_width=dp(8),
            scroll_type=["bars", "content"],
            size_hint=(1, None),
            height=row_h * 9,
        )
        sv.effect_cls = ScrollEffect
        # контент: вертикальный список строк
        rows = GridLayout(cols=1, spacing=0, size_hint_y=None)
        rows.bind(minimum_height=rows.setter("height"))
        sv.add_widget(rows)
        outer.add_widget(sv)
        pop = Popup(
            title="",
            content=outer,
            size_hint=(0.592, None),
            height=ti_h + row_h * 9,
            auto_dismiss=True,
            separator_height=0,
            title_size=0,
            padding=0,
            border=(0, 0, 0, 0),
            background="",          # убираем стандартную рамку/фон
            background_color=(0, 0, 0, 0),
            overlay_color=(0, 0, 0, 0.35),
        )
        # --- динамическая высота списка: чтобы результаты "прилипали" к полю поиска ---
        def _sync_popup_heights(*_args):
            content_h = rows.height
            max_h = row_h * 9
            if content_h > max_h:
                content_h = max_h
            if content_h < row_h:
                content_h = row_h
            sv.height = content_h
            outer.height = ti_h + content_h
            pop.height = outer.height
        rows.bind(height=_sync_popup_heights)
        _sync_popup_heights()
        # -------- заполнение списка --------
        def _shop_text(shop: dict) -> str:
            name = str(shop.get("name", "")).strip()
            phone = str(shop.get("phone", "")).strip()
            note = str(shop.get("note", "")).strip()
            parts = [p for p in [name, phone, note] if p]
            return " • ".join(parts) if parts else ""
        def _rebuild(*_):
            rows.clear_widgets()
            query = (ti.text or "").strip().lower()
            try:
                shops = list(getattr(self, "shops", []) or [])
            except Exception:
                shops = []
            if query:
                def _hay(s):
                    return _shop_text(s).lower()
                shops = [s for s in shops if query in _hay(s)]
            if not shops:
                empty = Label(text="Нічого не знайдено", size_hint_y=None, height=row_h, color=(1, 1, 1, 0.85))
                _apply_bg(empty, Z2)
                rows.add_widget(empty)
                return
            for i, shop in enumerate(shops):
                # row container with zebra bg
                row = BoxLayout(orientation="horizontal", size_hint_y=None, height=row_h, spacing=0, padding=[0, 0, 0, 0])
                _apply_bg(row, Z1 if i % 2 == 0 else Z2)
                # left text (дві строки: назва + телефон/адреса)
                name = str(shop.get("name","")).strip()
                phone = str(shop.get("phone","")).strip()
                note = str(shop.get("note","")).strip()
                left_box = BoxLayout(orientation="vertical", padding=(dp(12), dp(4)), spacing=dp(2))
                title_lbl = Label(
                    text=f"[b]{name}[/b]",
                    markup=True,
                    halign="left",
                    valign="middle",
                    color=(1,1,1,1),
                    size_hint_y=None,
                    height=dp(22)
                )
                title_lbl.bind(size=lambda inst,*_: setattr(inst,"text_size",(inst.width,None)))
                second = f"тел.: {phone}  |  адр.: {note}" if (phone or note) else ""
                sub_lbl = Label(
                    text=second,
                    halign="left",
                    valign="middle",
                    color=(1,1,1,0.85),
                    size_hint_y=None,
                    height=dp(20)
                )
                sub_lbl.bind(size=lambda inst,*_: setattr(inst,"text_size",(inst.width,None)))
                left_box.add_widget(title_lbl)
                left_box.add_widget(sub_lbl)
                row.add_widget(left_box)
                # actions area (слепленные кнопки)
                actions = BoxLayout(orientation="horizontal", size_hint_x=None, width=btn_col_w, spacing=0, padding=[0, 0, 0, 0])
                # icons: pencil / trash
                edit_icon = "\uE104" if ICON_FONT else "✎"
                del_icon = "\uE107" if ICON_FONT else "🗑"
                btn_edit = _style_icon_btn(Button(text=edit_icon, size_hint=(None, 1), width=btn_w), BTN_EDIT_UP, BTN_EDIT_DOWN)
                btn_del = _style_icon_btn(Button(text=del_icon, size_hint=(None, 1), width=btn_w), BTN_DEL_UP, BTN_DEL_DOWN)
                def _on_edit(_btn, _shop=shop):
                    try:
                        self._open_edit_shop_popup(_shop, on_saved=lambda *_: _rebuild())
                    except Exception:
                        pass
                def _on_del(_btn, _shop=shop):
                    try:
                        self._try_delete_shop(_shop, on_done=lambda *_: _rebuild())
                    except Exception as e:
                        self._notify('Помилка', str(e))
                btn_edit.bind(on_release=_on_edit)
                # also clear any press highlight when action buttons used
                def _clear_press(_btn, _row=row):
                    try:
                        _press_off(_row)
                    except Exception:
                        pass
                btn_edit.bind(on_press=_clear_press)
                btn_del.bind(on_press=_clear_press)
                btn_del.bind(on_release=_on_del)
                # Tap on the zebra row selects the shop (excluding action buttons)
                from time import time as _time
                def _row_on_down(_inst, _touch, _shop=shop, _be=btn_edit, _bd=btn_del):
                    if _inst.collide_point(*_touch.pos) and not _be.collide_point(*_touch.pos) and not _bd.collide_point(*_touch.pos):
                        _inst._tap = (_touch.x, _touch.y, _time())
                        try:
                            _press_on(_inst)
                        except Exception:
                            pass
                    return False
                def _row_on_up(_inst, _touch, _shop=shop, _be=btn_edit, _bd=btn_del):
                    tap = getattr(_inst, '_tap', None)
                    if tap and _inst.collide_point(*_touch.pos) and not _be.collide_point(*_touch.pos) and not _bd.collide_point(*_touch.pos):
                        x0, y0, t0 = tap
                        if abs(_touch.x - x0) < dp(12) and abs(_touch.y - y0) < dp(12) and (_time() - t0) < 0.45:
                            name = (_shop or {}).get('name', '') or ''
                            self.apply_shop_selection(name)
                            try:
                                pop.dismiss()
                            except Exception:
                                pass
                    _inst._tap = None
                    try:
                        _press_off(_inst)
                    except Exception:
                        pass
                    return False
                row.bind(on_touch_down=_row_on_down, on_touch_up=_row_on_up)
                def _row_on_move(_inst, _touch, _be=btn_edit, _bd=btn_del):
                    # if user drags/scrolls, remove highlight
                    if getattr(_inst, '_tap', None) and _inst.collide_point(*_touch.pos) and not _be.collide_point(*_touch.pos) and not _bd.collide_point(*_touch.pos):
                        x0, y0, _t0 = _inst._tap
                        if abs(_touch.x - x0) > dp(12) or abs(_touch.y - y0) > dp(12):
                            _inst._tap = None
                            try:
                                _press_off(_inst)
                            except Exception:
                                pass
                    return False
                row.bind(on_touch_move=_row_on_move)
                actions.add_widget(btn_edit)
                actions.add_widget(btn_del)
                row.add_widget(actions)
                rows.add_widget(row)
        # rebuild on open + on text change (deferred, чтобы не лагало)
        def _schedule_rebuild(*_):
            Clock.unschedule(_rebuild)
            Clock.schedule_once(_rebuild, 0)
        ti.bind(text=_schedule_rebuild)
        _rebuild()  # build before open to avoid staged rendering
        pop.open()
        def _clear_focus(*_):
            for w in (getattr(self, '_openinv_ti_q', None), getattr(self, '_openinv_ti_from', None), getattr(self, '_openinv_ti_to', None)):
                if w is not None:
                    try:
                        w.focus = False
                    except Exception:
                        pass
        Clock.schedule_once(_clear_focus, 0)
    def _delete_shop_and_refresh(self, shop, refresh_cb):
        # безопасное удаление (чтобы popup не падал) + перерисовка
        try:
            # ожидаем, что shops_data — список dict'ов
            if isinstance(self.shops_data, list):
                try:
                    self.shops_data.remove(shop)
                except ValueError:
                    # если объекты не совпали, попробуем по id/имени
                    sid = shop.get("id", None) if isinstance(shop, dict) else None
                    if sid is not None:
                        self.shops_data = [s for s in self.shops_data if s.get("id", None) != sid]
                    else:
                        name = shop.get("name", "")
                        addr = shop.get("address", "")
                        self.shops_data = [s for s in self.shops_data if not (s.get("name", "") == name and s.get("address", "") == addr)]
            # если есть функция сохранения — вызываем
            if hasattr(self, "save_shops_data"):
                try:
                    self.save_shops_data()
                except Exception:
                    pass
        finally:
            try:
                refresh_cb()
            except Exception:
                pass
    def open_edit_shop_popup(self, shop: dict, on_saved=None):
        # публичный алиас (для совместимости)
        return self._open_edit_shop_popup(shop, on_saved=on_saved)
    def _open_edit_shop_popup(self, shop: dict = None, on_saved=None, is_new: bool = False):
        # shop is a dict from self.shops
        from kivy.uix.popup import Popup
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.textinput import TextInput
        from kivy.uix.button import Button
        from kivy.uix.label import Label
        from kivy.metrics import dp, sp
        from kivy.clock import Clock

        if shop is None:
            shop = {}
        old_name = str(shop.get("name", "")).strip()
        # match manage popup sizes
        row_h = dp(54)
        ti_h = dp(36)
        # same visual as TopBtn
        BTN_UP = (0.30, 0.30, 0.30, 1)
        BTN_DOWN = (0.1529, 0.6784, 0.9608, 1)
        def _style_btn(btn: Button):
            btn.background_normal = ""
            btn.background_down = ""
            btn.color = (1, 1, 1, 1)
            btn._bg_up = BTN_UP
            btn._bg_down = BTN_DOWN
            btn.background_color = btn._bg_up
            def _on_state(_btn, _state):
                _btn.background_color = _btn._bg_down if _state == "down" else _btn._bg_up
            btn.bind(state=_on_state)
            return btn
        root_box = BoxLayout(orientation="vertical", spacing=dp(10), padding=[dp(12), dp(12), dp(12), dp(12)])
        ti_name = TextInput(
            text=old_name,
            hint_text="Назва магазину",
            multiline=False,
            font_size=dp(18),
            padding=[dp(10), dp(6), dp(10), dp(6)],
            size_hint_y=None,
            height=ti_h,
            background_normal="",
            background_active="",
            background_color=(1, 1, 1, 1),
            foreground_color=(0, 0, 0, 1),
            hint_text_color=(0.65, 0.65, 0.65, 1),
            cursor_color=(0, 0, 0, 1),
        )
        # Prefill quick-add name from what user typed in the main shop field
        if is_new and not old_name:
            try:
                pending = (getattr(self, "_pending_shop_text", "") or "").strip()
            except Exception:
                pending = ""
            if pending:
                ti_name.text = pending
        ti_phone = TextInput(
            text=str(shop.get("phone", "")).strip(),
            hint_text="Телефон (необов'язково)",
            multiline=False,
            font_size=dp(18),
            padding=[dp(10), dp(6), dp(10), dp(6)],
            size_hint_y=None,
            height=ti_h,
            background_normal="",
            background_active="",
            background_color=(1, 1, 1, 1),
            foreground_color=(0, 0, 0, 1),
            hint_text_color=(0.65, 0.65, 0.65, 1),
            cursor_color=(0, 0, 0, 1),
        )
        ti_note = TextInput(
            text=str(shop.get("note", "")).strip(),
            hint_text="Коментар/адреса (необов'язково)",
            multiline=False,
            font_size=dp(18),
            padding=[dp(10), dp(6), dp(10), dp(6)],
            size_hint_y=None,
            height=ti_h,
            background_normal="",
            background_active="",
            background_color=(1, 1, 1, 1),
            foreground_color=(0, 0, 0, 1),
            hint_text_color=(0.65, 0.65, 0.65, 1),
            cursor_color=(0, 0, 0, 1),
        )
        msg = Label(text="", size_hint_y=None, height=dp(22), color=(1, 0, 0, 1))
        btn_row = BoxLayout(size_hint_y=None, height=row_h, spacing=dp(10))
        btn_cancel = _style_btn(Button(text="Скасувати", font_size=dp(18)))
        btn_save = _style_btn(Button(text="Зберегти", font_size=dp(18)))
        btn_row.add_widget(btn_cancel)
        btn_row.add_widget(btn_save)
        root_box.add_widget(ti_name)
        root_box.add_widget(ti_phone)
        root_box.add_widget(ti_note)
        root_box.add_widget(msg)
        root_box.add_widget(btn_row)
        pop = Popup(title=("Додати магазин" if is_new else "Редагувати магазин"),
            content=root_box,
            size_hint=(None, None),
            size=(dp(520), dp(320)),
            auto_dismiss=False,
        )
        
        def _do_cancel(*_):
            pop.dismiss()

        def _do_save(*_):
            name = (ti_name.text or "").strip()
            if not name:
                msg.text = "Вкажіть назву магазину"
                return

            if is_new:
                if self._shop_name_exists(name):
                    msg.text = "Такий магазин вже існує"
                    return
                new_shop = {
                    "name": name,
                    "phone": (ti_phone.text or "").strip(),
                    "note": (ti_note.text or "").strip(),
                }
                self.shops.append(new_shop)
                try:
                    self.shops.sort(key=lambda s: str(s.get("name", "")).lower())
                except Exception:
                    pass
                self._shops_file_write()

                self.selected_shop = name
                if "shop" in self.ids:
                    self.ids.shop.text = name
                self._save_last_shop(name)
            else:
                if self._shop_name_exists(name, exclude_name=old_name):
                    msg.text = "Такий магазин вже існує"
                    return
                shop["name"] = name
                shop["phone"] = (ti_phone.text or "").strip()
                shop["note"] = (ti_note.text or "").strip()
                try:
                    self.shops.sort(key=lambda s: str(s.get("name", "")).lower())
                except Exception:
                    pass
                self._shops_file_write()

                if (self.selected_shop or "").strip() == old_name:
                    self.selected_shop = name
                    if "shop" in self.ids:
                        self.ids.shop.text = name
                    self._save_last_shop(name)

            pop.dismiss()
            if callable(on_saved):
                on_saved()

        btn_cancel.bind(on_release=_do_cancel)
        btn_save.bind(on_release=_do_save)
        pop.open()
        def _focus_name(_dt):
            try:
                ti_name.focus = True
                ti_name.select_all()
            except Exception:
                pass
        Clock.schedule_once(_focus_name, 0)
    def _try_delete_shop(self, shop, on_done=None):
        # "Видалити" в нас означає: прибрати зі списку + перенести папку магазину в "Архів" (якщо вона існує)
        if not shop:
            return
        name = (shop.get("name") or "").strip()
        if not name:
            name = "Магазин"
        def _yes(*_):
            # 1) Перенести папку рахунків магазину в Архів (якщо є)
            try:
                base_dir = self._storage_base_dir()
                shop_dir = os.path.join(base_dir, self._safe_folder_name(name))
                archive_root = os.path.join(base_dir, "Архів")
                os.makedirs(archive_root, exist_ok=True)
                if os.path.isdir(shop_dir):
                    target = os.path.join(archive_root, os.path.basename(shop_dir))
                    # якщо в архіві вже є така папка — додаємо суфікс з датою/часом
                    if os.path.exists(target):
                        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        target = os.path.join(archive_root, f"{os.path.basename(shop_dir)}__{ts}")
                    shutil.move(shop_dir, target)
            except Exception as e:
                Logger.warning(f"ARCHIVE: Не вдалося перенести папку магазину в Архів: {e}")
            # 2) Видалення магазину зі списку (shops.json)
            try:
                self.shops = [s for s in self.shops if s is not shop]
            except Exception:
                # fallback: видаляємо за полями
                self.shops = [s for s in self.shops if (s.get("name"), s.get("phone"), s.get("note")) != (shop.get("name"), shop.get("phone"), shop.get("note"))]
            self._shops_file_write()
            self._refresh_shops_list()
            self._notify("Готово", f"Магазин перенесено в Архів і прибрано зі списку:\n{name}")
            if callable(on_done):
                try:
                    on_done()
                except Exception:
                    pass
        self._confirm(
            "Видалення магазину",
            f"Перенести магазин в папку Архів?\n\n[b]{name}[/b]",
            on_yes=_yes,
        )
    def on_shop_focus(self, focused: bool):
        if focused:
            self.on_shop_text(self.ids.shop.text if "shop" in self.ids else "")
        else:
            self.dismiss_shop_dropdown()
    def on_shop_text(self, text: str):
        # Debounce to next frame to avoid focus/glitch when there are matches
        self._pending_shop_text = text or ""
        Clock.unschedule(self._shop_dropdown_scheduled)
        Clock.schedule_once(self._shop_dropdown_scheduled, 0)
    def _shop_dropdown_scheduled(self, *_):
        self._show_shop_dropdown(self._pending_shop_text)
    def dismiss_shop_dropdown(self):
        if self._shop_dd is not None:
            try:
                self._shop_dd.dismiss()
            except Exception:
                pass
    def _find_shop_matches(self, text: str, limit: int = 5000):
        q = (text or "").strip().lower()
        if not self.shops:
            return []
        if not q:
            return list(range(min(limit, len(self.shops))))
        out = []
        for i, s in enumerate(self.shops):
            hay = " ".join([
                str(s.get("name","")),
                str(s.get("phone","")),
                str(s.get("note","")),
            ]).lower()
            if q in hay:
                out.append(i)
                if len(out) >= limit:
                    break
        return out
    def _show_shop_dropdown(self, text: str):
        if "shop" not in self.ids:
            return
        inp = self.ids.shop
        if not inp.focus:
            return
        inds = self._find_shop_matches(text, limit=5000)
        if not inds:
            self.dismiss_shop_dropdown()
            return
        from kivy.uix.dropdown import DropDown
        class _DDShopItem(Button):
            def __init__(self, base_rgba, **kwargs):
                super().__init__(**kwargs)
                self._base_rgba = base_rgba
                self.background_normal = ""
                self.background_down = ""
                self.background_color = base_rgba
                self.markup = True
                self.halign = "left"
                self.valign = "middle"
                self.padding = (dp(12), dp(6))
                self.bind(state=self._on_state)
            def _on_state(self, _inst, state):
                if state == "down":
                    self.background_color = (0.1529, 0.6784, 0.9608, 1)
                else:
                    self.background_color = self._base_rgba
        if self._shop_dd is None:
            self._shop_dd = DropDown(auto_width=False, max_height=dp(9 * 66 + 16))
            # Make dropdown scrolling feel like the popup list (no rubber-band overscroll)
            try:
                from kivy.effects.scroll import ScrollEffect
                self._shop_dd.effect_cls = ScrollEffect
                self._shop_dd.scroll_type = ["bars", "content"]
                self._shop_dd.do_scroll_x = False
                self._shop_dd.bar_width = dp(8)
            except Exception:
                pass
            # Visible rows like popup: 9 * row_h (row_h=54dp)
            try:
                self._shop_dd.max_height = dp(54 * 9)
            except Exception:
                pass
        dd = self._shop_dd
        dd.clear_widgets()
        dd.width = inp.width
        for pos, real_idx in enumerate(inds):
            s = self.shops[real_idx]
            title = str(s.get("name", "")).strip()
            phone_raw = str(s.get("phone", "")).strip()
            note = str(s.get("note", "")).strip()
            # нормалізація телефонів (може бути 2, через кому)
            phones = []
            if phone_raw:
                for part in phone_raw.split(","):
                    p = part.strip()
                    if p:
                        phones.append(p)
            phone_txt = ", ".join(phones)
            # 2 рядки максимум:
            #   1) Назва (жирним)
            #   2) тел/адр через " | " + обрізка з "..." якщо не влазить
            meta_parts = []
            if phone_txt:
                meta_parts.append(f"тел.: {phone_txt}")
            if note:
                meta_parts.append(f"адр.: {note}")
            second = " | ".join(meta_parts).strip()
            indent = ""
            max_title = 60
            max_second = 80
            t = title
            if len(t) > max_title:
                t = t[: max_title - 3].rstrip() + "..."
            if len(second) > max_second:
                second = second[: max_second - 3].rstrip() + "..."
            line = f"[b]{indent}{t}[/b]" + (f"\n{indent}{second}" if second else "")
                        # Zebra colors like popup
            Z1 = (0.38, 0.38, 0.38, 1)
            Z2 = (0.44, 0.44, 0.44, 1)
            # Two-line markup: bold title + lighter meta
            line = f"[b]{t}[/b]" + (f"\n[color=dddddd]{second}[/color]" if second else "")
            btn = _DDShopItem(
                base_rgba=(Z1 if (pos % 2 == 0) else Z2),
                text=line,
                size_hint_y=None,
                height=dp(54),
            )
            btn.bind(size=lambda inst, val: setattr(inst, "text_size", (val[0], None)))
            def _select(_btn, name=title):
                self.selected_shop = name
                inp.text = name
                self._save_last_shop(name)
                self.dismiss_shop_dropdown()
                Clock.schedule_once(lambda dt: setattr(inp, "focus", True), 0)
            btn.bind(on_release=_select)
            dd.add_widget(btn)
        # open anchored to input (doesn't steal focus like Popup)
        # IMPORTANT: don't call open() repeatedly while typing
        if getattr(dd, "attach_to", None) is None:
            dd.open(inp)
    # --- Top bar actions (save / open / print) ---
    def _cart_save_path(self) -> str:
        """Default path for cart save file (next to this script)."""
        try:
            base = app_dir()
        except Exception:
            base = os.getcwd()
        return os.path.join(base, "cart_saved.json")
    def _notify(self, title: str, message: str) -> None:
        """Simple notification popup.
        Uses a ScrollView so long messages (like file paths) don't overflow.
        """
        try:
            root = BoxLayout(orientation="vertical", padding=dp(10), spacing=dp(10))
            sv = ScrollView(do_scroll_x=False, do_scroll_y=True)
            msg = Label(
                text=message,
                size_hint_y=None,
                halign="left",
                valign="top",
            )
            # Make text wrap to popup width and expand height as needed
            msg.bind(width=lambda inst, w: setattr(inst, "text_size", (w, None)))
            msg.bind(texture_size=lambda inst, ts: setattr(inst, "height", ts[1]))
            sv.add_widget(msg)
            root.add_widget(sv)
            # Optional close button (still auto_dismiss on outside tap)
            # OK button row (fixed height, like confirm popup buttons)
            btn_row = BoxLayout(size_hint=(1, None), height=dp(62))
            btn = PopupButton(text="OK", size_hint=(1, 1))
            btn_row.add_widget(btn)
            root.add_widget(btn_row)
            p = Popup(title=title, content=root, size_hint=(None, None), size=(dp(442), dp(204)), auto_dismiss=True)
            btn.bind(on_release=lambda *_: p.dismiss())
            p.open()
        except Exception:
            print(f"[{title}] {message}")
    
    def _notify_then(self, title: str, message: str, after_ok=None) -> None:
        """Notification popup that runs after_ok only after user presses OK."""
        try:
            root = BoxLayout(orientation="vertical", padding=dp(10), spacing=dp(10))
            sv = ScrollView(do_scroll_x=False, do_scroll_y=True)
            msg = Label(text=message, size_hint_y=None, halign="left", valign="top")
            msg.bind(width=lambda inst, w: setattr(inst, "text_size", (w, None)))
            msg.bind(texture_size=lambda inst, ts: setattr(inst, "height", ts[1]))
            sv.add_widget(msg)
            root.add_widget(sv)

            # OK button row (fixed height, like confirm popup buttons)
            btn_row = BoxLayout(size_hint=(1, None), height=dp(62))
            btn = PopupButton(text="OK", size_hint=(1, 1))
            btn_row.add_widget(btn)
            root.add_widget(btn_row)

            p = Popup(title=title, content=root, size_hint=(None, None), size=(dp(442), dp(204)), auto_dismiss=True)

            def _close_and_then(*_):
                try:
                    p.dismiss()
                finally:
                    if callable(after_ok):
                        after_ok()

            btn.bind(on_release=_close_and_then)
            p.open()
        except Exception:
            print(f"[{title}] {message}")
            if callable(after_ok):
                after_ok()
    def _confirm_yes_no_simple(self, title: str, message: str, on_yes, on_no) -> None:
        """Popup with Так/Ні. Dismiss by tapping outside. Calls callbacks as-is."""
        box = BoxLayout(orientation='vertical', spacing=dp(10), padding=dp(10))
        box.add_widget(Label(text=message, markup=True))

        btns = BoxLayout(size_hint_y=None, height=dp(62), spacing=dp(10))
        b_yes = PopupButton(text='Так')
        b_no = PopupButton(text='Ні')
        b_yes.size_hint_x = 1
        b_no.size_hint_x = 1

        btns.add_widget(b_yes)
        btns.add_widget(b_no)
        box.add_widget(btns)

        pop = Popup(title=title, content=box, size_hint=(None, None), size=(dp(442), dp(204)), auto_dismiss=True)

        def _yes(*_):
            pop.dismiss()
            try:
                if callable(on_yes):
                    on_yes()
            except Exception as e:
                self._notify('Помилка', str(e))

        def _no(*_):
            pop.dismiss()
            try:
                if callable(on_no):
                    on_no()
            except Exception as e:
                self._notify('Помилка', str(e))

        b_yes.bind(on_release=_yes)
        b_no.bind(on_release=_no)
        pop.open()

    def _confirm(self, title: str, message: str, on_yes=None, on_no=None):
        return self._confirm_yes_no_simple(title, message, on_yes, on_no)
    def _clear_cart(self) -> None:
        self.cart.clear()
        self.edit_article = ""
        self.edit_field = ""
        try:
            if 'order_comment' in self.ids:
                self.ids.order_comment.text = ''
        except Exception:
            pass
        self.refresh_cart()
        self._cart_dirty = False
    def new_account(self) -> None:
        """Новый рахунок / очистка текущего счёта (нижняя кнопка)."""
        if not self.cart:
            return
        if getattr(self, '_cart_dirty', False):
            self._confirm_yes_no_simple(
                title='Підтвердження',
                message='Рахунок не пустий. Зберегти його?',
                on_yes=lambda: (self.save_cart() and self._clear_cart()),
                on_no=self._clear_cart,
            )
        else:
            self._clear_cart()
    def save_cart(self, show_popup: bool = True) -> bool:
        """Save current cart to an Excel file inside a folder named by shop."""
        # If cart is empty -> do nothing
        if not self.cart:
            return False
        shop_name = (self.ids.shop.text if hasattr(self.ids, 'shop') else '').strip()
        if not shop_name:
            self._notify("Увага", "Створіть або виберіть магазин.")
            return False
        # If user typed a new shop manually, add it to the shops list immediately
        self._ensure_shop_in_list(shop_name)
        self._save_last_shop(shop_name)
        try:
            path = self._save_cart_to_excel(shop_name)
        except Exception as e:
            self._notify("Помилка", f"Не вдалося зберегти рахунок:\n{e}")
            return False
        if show_popup:
            self._notify("Збережено", f"Рахунок збережено у файл:\n{path}")
        self._cart_dirty = False
        self._cart_last_saved_path = str(path)
        return True

    def save_cart_silent(self):
        """Save cart without showing a popup. Returns saved path (str) on success, else None."""
        if not self.cart:
            return None
        shop_name = (self.ids.shop.text if hasattr(self.ids, 'shop') else '').strip()
        if not shop_name:
            # keep same behavior as before (user must pick shop)
            self._notify("Увага", "Створіть або виберіть магазин.")
            return None
        # If user typed a new shop manually, add it to the shops list immediately
        self._ensure_shop_in_list(shop_name)
        self._save_last_shop(shop_name)
        try:
            path = self._save_cart_to_excel(shop_name)
        except Exception as e:
            self._notify("Помилка", f"Не вдалося зберегти рахунок:\n{e}")
            return None
        self._cart_dirty = False
        self._cart_last_saved_path = str(path)
        return str(path)
    def open_cart(self) -> None:
        """Open (load) a cart from Excel. If current cart is not empty, ask to save it."""
        if self.cart and getattr(self, '_cart_dirty', False):
            self._confirm_yes_no(
                title="Підтвердження",
                message="Рахунок не пустий. Зберегти його?",
                on_yes=self.save_cart_silent,
                on_no=self._open_cart_filechooser,
            )
        else:
            self._open_cart_filechooser()
    def _confirm_yes_no(self, title: str, message: str, on_yes, on_no) -> None:
        """Confirm before opening рахунок file chooser. Reuses the same popup style.

        If on_yes saves the рахунок, the file chooser will open ONLY after user closes
        the "Збережено" popup (requested UX).
        """

        def _yes_action():
            result = None
            try:
                result = on_yes() if callable(on_yes) else None
            except Exception as e:
                self._notify('Помилка', str(e))
                result = None

            if result:
                # result may be True or a path string
                path = result if isinstance(result, str) else getattr(self, '_cart_last_saved_path', '')
                # Show saved popup first, then open chooser after OK
                self._notify_then('Збережено', f'Рахунок збережено у файл:\n{path}' if path else 'Рахунок збережено.', after_ok=self._open_cart_filechooser)
            else:
                # If not saved (or user has empty cart), don't proceed automatically
                # (keeps previous safety: only open chooser if save succeeded)
                pass

        def _no_action():
            try:
                if callable(on_no):
                    on_no()
            except Exception as e:
                self._notify('Помилка', str(e))

        return self._confirm_yes_no_simple(title, message, _yes_action, _no_action)

    def _open_cart_filechooser(self) -> None:
        """Fast popup to pick an Excel рахунок from рахунки/<shop>/*.xlsx with search + date filter."""
        base = self._storage_base_dir()
        os.makedirs(base, exist_ok=True)

        # --- scan рахунки folders (shop subfolders) ---
        rows = []
        try:
            for root, dirs, files in os.walk(base):
                # skip archive folders
                dirs[:] = [d for d in dirs if d.lower() not in ('архів','архив','archive')]
                for fn in files:
                    if not fn.lower().endswith(".xlsx"):
                        continue
                    p = os.path.join(root, fn)
                    try:
                        mtime = os.path.getmtime(p)
                    except Exception:
                        mtime = 0
                    dt = datetime.datetime.fromtimestamp(mtime) if mtime else datetime.datetime.fromtimestamp(0)
                    shop = os.path.basename(os.path.dirname(p))
                    # if file is directly in base, shop is empty -> show as "-"
                    if os.path.abspath(os.path.dirname(p)) == os.path.abspath(base):
                        shop = "-"
                    rows.append({
                        "path": p,
                        "shop": shop,
                        "fname": fn,
                        "mdate": dt.date(),
                        "mtime_text": dt.strftime("%Y-%m-%d %H:%M"),
                    })
        except Exception as e:
            self._notify("Помилка", f"Не вдалося прочитати папку рахунків:\n{e}")
            return

        rows.sort(key=lambda x: x.get("mtime_text", ""), reverse=True)

        # --- UI ---
        root = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(6), dp(6), dp(6), dp(0)])

        # Filters row (flat style like main window)
        row_h = dp(36)
        filters = BoxLayout(size_hint_y=None, height=row_h, spacing=dp(6))

        def _make_top_ti(hint: str, width_hint: float):
            # Поля як на головному вікні: height 36dp, font 18dp, плоскі, без "стрибків" при фокусі
            ti = TextInput(
                hint_text=hint,
                hint_text_color=(0.55, 0.55, 0.55, 1),
                multiline=False,
                size_hint_y=None,
                height=dp(36),
                size_hint_x=width_hint,
                font_size=dp(18),
                background_normal="",
                background_active="",
                background_color=(1, 1, 1, 1),
                foreground_color=(0, 0, 0, 1),
                cursor_color=(0, 0, 0, 1),
                cursor_width=dp(2),
                cursor_blink=False,
                # padding: [left, top, right, bottom]
                padding=[dp(10), dp(7), dp(10), dp(7)],
                write_tab=False,
            )
            return ti

        def _make_top_btn(text: str, width_hint: float):
            b = PopupButton(text=text)
            b.size_hint_x = width_hint
            b.size_hint_y = None
            b.height = dp(36)
            b.font_size = dp(18)
            return b

        ti_q = _make_top_ti("Пошук (магазин/файл)", 0.44)
        ti_from = _make_top_ti("Від (YYYY-MM-DD)", 0.17)
        ti_to = _make_top_ti("До (YYYY-MM-DD)", 0.17)
        # keep refs for safe focus clearing
        self._openinv_ti_q = ti_q
        self._openinv_ti_from = ti_from
        self._openinv_ti_to = ti_to
        # make sure hints are visible on open
        ti_q.text = ""
        ti_from.text = ""
        ti_to.text = ""


        btn_7 = _make_top_btn("7д", 0.05)
        btn_30 = _make_top_btn("30д", 0.06)
        btn_clear = _make_top_btn("Очист.", 0.08)
        btn_explorer = _make_top_btn("Провідник", 0.13)

        filters.add_widget(ti_q)
        filters.add_widget(ti_from)
        filters.add_widget(ti_to)
        filters.add_widget(btn_7)
        filters.add_widget(btn_30)
        filters.add_widget(btn_clear)
        filters.add_widget(btn_explorer)

        root.add_widget(filters)

        rv = InvoicesRV()
        rv.set_all(rows)
        root.add_widget(rv)

        # Bottom buttons
        bar = BoxLayout(size_hint_y=None, height=dp(54), spacing=dp(6))
        btn_open = PopupButton(text="Відкрити")
        btn_cancel = PopupButton(text="Закрити")
        bar.add_widget(Widget())
        bar.add_widget(btn_cancel)
        bar.add_widget(btn_open)
        root.add_widget(bar)

        pop = Popup(title="Відкрити рахунок", content=root, size_hint=(0.75, 0.75), auto_dismiss=True)

        def apply_filters(*_):
            rv.apply_filter(ti_q.text, ti_from.text, ti_to.text)

        def quick_days(days: int):
            today = datetime.date.today()
            d_from = today - datetime.timedelta(days=int(days))
            ti_from.text = d_from.strftime("%Y-%m-%d")
            ti_to.text = today.strftime("%Y-%m-%d")
            apply_filters()

        def clear_filters(*_):
            ti_q.text = ""
            ti_from.text = ""
            ti_to.text = ""
            rv.apply_filter("", "", "")

        # Debounced typing for search
        def on_q_change(_inst, _val):
            Clock.unschedule(apply_filters)
            Clock.schedule_once(apply_filters, 0.15)

        ti_q.bind(text=on_q_change)
        ti_from.bind(text=on_q_change)
        ti_to.bind(text=on_q_change)

        btn_7.bind(on_release=lambda *_: quick_days(7))
        btn_30.bind(on_release=lambda *_: quick_days(30))
        btn_clear.bind(on_release=clear_filters)

        # --- native file explorer picker (optional) ---
        def _on_explorer_selection(selection):
            if not selection:
                return
            p = selection[0]
            if not isinstance(p, str) or not p.lower().endswith(".xlsx"):
                self._notify("Увага", "Вибери файл .xlsx")
                return
            pop.dismiss()
            try:
                self._load_cart_from_excel(p)
                self.refresh_cart()
            except Exception as e:
                self._notify("Помилка", f"Не вдалося відкрити файл:\n{e}")

        def open_in_explorer(*_):
            """Відкрити через простий список-провідник (без plyer)."""
            fc = FileChooserListView(
                path=base,
                filters=["*.xlsx"],
                show_hidden=False,
                multiselect=False,
                dirselect=False,
            )
            # чуть крупнее строки, чтобы было как в приложении
            try:
                fc.font_size = dp(18)
            except Exception:
                pass

            root2 = BoxLayout(orientation="vertical", spacing=dp(6), padding=[dp(8), dp(8), dp(8), dp(8)])
            root2.add_widget(fc)

            bar2 = BoxLayout(size_hint_y=None, height=dp(54), spacing=dp(6))
            b_cancel2 = PopupButton(text="Закрити")
            b_open2 = PopupButton(text="Відкрити")
            bar2.add_widget(Widget())
            bar2.add_widget(b_cancel2)
            bar2.add_widget(b_open2)
            root2.add_widget(bar2)

            pop2 = Popup(title="Провідник", content=root2, size_hint=(0.66, 0.66), auto_dismiss=True)

            def _do_open2(*__):
                sel = getattr(fc, "selection", []) or []
                if not sel:
                    self._notify("Інфо", "Оберіть файл .xlsx")
                    return
                p = sel[0]
                if not isinstance(p, str) or not p.lower().endswith(".xlsx"):
                    self._notify("Увага", "Вибери файл .xlsx")
                    return
                try:
                    self._load_cart_from_excel(p)
                    self.refresh_cart()
                    pop2.dismiss()
                    pop.dismiss()
                except Exception as e:
                    self._notify("Помилка", f"Не вдалося відкрити файл:\n{e}")

            b_cancel2.bind(on_release=lambda *_: pop2.dismiss())
            b_open2.bind(on_release=_do_open2)
            pop2.open()


        btn_explorer.bind(on_release=open_in_explorer)


        def do_open(*_):
            p = rv.selected_path
            if not p:
                self._notify("Увага", "Вибери файл зі списку.")
                return
            pop.dismiss()
            try:
                self._load_cart_from_excel(p)
                self.refresh_cart()
            except Exception as e:
                self._notify("Помилка", f"Не вдалося відкрити файл:\n{e}")

        btn_open.bind(on_release=do_open)
        btn_cancel.bind(on_release=lambda *_: pop.dismiss())

        pop.open()
    def _storage_base_dir(self) -> str:
        """Base directory where shop folders will be stored."""
        try:
            base = app_dir()
        except Exception:
            base = os.getcwd()
        return os.path.join(base, 'рахунки')
    def _safe_folder_name(self, name: str, max_len: int = 50) -> str:
        """Make a filesystem-safe folder name (Windows-friendly)."""
        name = (name or '').strip()
        # replace forbidden characters for Windows and most filesystems
        name = re.sub(r'[\\/:*?"<>|]+', '_', name)
        # remove control chars
        name = ''.join(ch for ch in name if ord(ch) >= 32)
        # collapse whitespace
        name = re.sub(r'\s+', ' ', name).strip()
        # Windows doesn't like trailing dot/space
        name = name.rstrip(' .')
        if not name:
            name = 'Магазин'
        if len(name) > max_len:
            name = name[:max_len].rstrip(' .')
        return name
    def _save_cart_to_excel(self, shop_name: str) -> str:
        """
        Save current cart to Excel with printable A4 portrait layout (based on older template).
        We keep logic/data from current app, but layout matches the "Замовлення клієнта ..." style.
        """
        safe_shop = self._safe_folder_name(shop_name)
        folder = os.path.join(self._storage_base_dir(), safe_shop)
        os.makedirs(folder, exist_ok=True)
        dt = datetime.datetime.now()
        inv_id = dt.strftime('%Y%m%d%H%M%S')
        filename = dt.strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'
        path = os.path.join(folder, filename)

        # ---- helpers (local, so we don't touch the rest of the app) ----
        import math
        def money2(x: float) -> float:
            try:
                return round(float(x) + 1e-9, 2)
            except Exception:
                return 0.0

        def _uk_choose_form(n: int, form1: str, form2: str, form5: str) -> str:
            n = abs(int(n))
            n10 = n % 10
            n100 = n % 100
            if 11 <= n100 <= 14:
                return form5
            if n10 == 1:
                return form1
            if 2 <= n10 <= 4:
                return form2
            return form5

        def _uk_triplet_to_words(num: int, gender: str) -> str:
            ones_m = ["", "один", "два", "три", "чотири", "п'ять", "шість", "сім", "вісім", "дев'ять"]
            ones_f = ["", "одна", "дві", "три", "чотири", "п'ять", "шість", "сім", "вісім", "дев'ять"]
            ones = ones_f if gender == "f" else ones_m
            teens = ["десять", "одинадцять", "дванадцять", "тринадцять", "чотирнадцять",
                     "п'ятнадцять", "шістнадцять", "сімнадцять", "вісімнадцять", "дев'ятнадцять"]
            tens = ["", "", "двадцять", "тридцять", "сорок", "п'ятдесят",
                    "шістдесят", "сімдесят", "вісімдесят", "дев'яносто"]
            hundreds = ["", "сто", "двісті", "триста", "чотириста", "п'ятсот",
                        "шістсот", "сімсот", "вісімсот", "дев'ятсот"]
            n = int(num)
            if n == 0:
                return ""
            parts = []
            h = n // 100
            t = (n // 10) % 10
            o = n % 10
            if h:
                parts.append(hundreds[h])
            if t == 1:
                parts.append(teens[o])
            else:
                if t:
                    parts.append(tens[t])
                if o:
                    parts.append(ones[o])
            return " ".join([p for p in parts if p]).strip()

        def uah_amount_to_words_uk(amount: float) -> str:
            try:
                amount = float(amount)
            except Exception:
                amount = 0.0
            amount = round(amount + 1e-9, 2)
            hryvnias = int(amount)
            kopecks = int(round((amount - hryvnias) * 100))
            if kopecks == 100:
                hryvnias += 1
                kopecks = 0

            scales = [
                ("", "", "", "m"),
                ("тисяча", "тисячі", "тисяч", "f"),
                ("мільйон", "мільйони", "мільйонів", "m"),
                ("мільярд", "мільярди", "мільярдів", "m"),
            ]

            if hryvnias == 0:
                words = "нуль"
            else:
                words_parts = []
                n = hryvnias
                scale_idx = 0
                while n > 0 and scale_idx < len(scales):
                    trip = n % 1000
                    if trip:
                        s1, s2, s5, g = scales[scale_idx]
                        trip_words = _uk_triplet_to_words(trip, g)
                        if scale_idx == 0:
                            words_parts.append(trip_words)
                        else:
                            scale_word = _uk_choose_form(trip, s1, s2, s5)
                            words_parts.append((trip_words + " " + scale_word).strip())
                    n //= 1000
                    scale_idx += 1
                words = " ".join(reversed([p for p in words_parts if p])).strip()

            грн_word = _uk_choose_form(hryvnias, "гривня", "гривні", "гривень")
            коп_word = _uk_choose_form(kopecks, "копійка", "копійки", "копійок")
            out = f"{words} {грн_word} {kopecks:02d} {коп_word}"
            return out[:1].upper() + out[1:]

        def format_date_uk(dt_obj) -> str:
            months = {
                1: "січня", 2: "лютого", 3: "березня", 4: "квітня",
                5: "травня", 6: "червня", 7: "липня", 8: "серпня",
                9: "вересня", 10: "жовтня", 11: "листопада", 12: "грудня",
            }
            try:
                return f"{int(dt_obj.day)} {months.get(int(dt_obj.month), '')} {int(dt_obj.year)} р."
            except Exception:
                return ""

        def estimate_lines(s: str, chars_per_line: int = 52) -> int:
            if not s:
                return 1
            s = str(s)
            parts = s.split("\n")
            lines = 0
            for p in parts:
                p = p.strip()
                if not p:
                    lines += 1
                    continue
                lines += max(1, int(math.ceil(len(p) / float(chars_per_line))))
            return max(1, lines)

        # try get phone/address/comment from shop
        shop_phone = ""
        shop_addr = ""
        try:
            for s in (self.shops or []):
                if str(s.get("name", "")).strip().lower() == str(shop_name).strip().lower():
                    shop_phone = str(s.get("phone", "") or "").strip()
                    shop_addr = str(s.get("note", "") or "").strip()  # in UI it's "Коментар/адреса"
                    break
        except Exception:
            shop_phone = ""
            shop_addr = ""

        # Build comments line (phone + address). Keep it compact and printable.
        parts = []
        if shop_phone:
            parts.append(f"тел.: {shop_phone}")
        if shop_addr:
            parts.append(f"адр.: {shop_addr}")
        shop_note = " | ".join(parts)

        wb = Workbook()
        ws = wb.active
        ws.title = "Рахунок"

        # styles
        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)
        thin = Side(style="thin", color="000000")
        thick = Side(style="thick", color="000000")
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill_header = PatternFill("solid", fgColor="E6E6E6")

        # clean sheet
        try:
            ws.sheet_view.showGridLines = False
        except Exception:
            pass

        # column widths (A4 portrait friendly)
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 54
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12

        # ---- Title ----
        ws.merge_cells("A1:F1")
        ws["A1"].value = f"Замовлення клієнта № {inv_id}     від {format_date_uk(dt)}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24
        for c in range(1, 7):
            ws.cell(row=1, column=c).border = Border(bottom=thick)

        # ---- Header fields (no box borders) ----
        ws.merge_cells("A2:B2")
        ws["A2"].value = "Виконавець:"
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("C2:F2")
        ws["C2"].value = "BRADAS"
        ws["C2"].font = bold
        ws["C2"].alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A3:B3")
        ws["A3"].value = "Замовник:"
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("C3:F3")
        ws["C3"].value = str(shop_name)
        ws["C3"].font = bold
        ws["C3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 18

        ws.merge_cells("A4:B4")
        ws["A4"].value = "Контакти:"
        ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("C4:F4")
        ws["C4"].value = shop_note
        ws["C4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[4].height = 30 if shop_note else 18

        # Рядок "Коментарі" з програми (під контактами)
        order_comment = ""
        try:
            order_comment = str(self.ids.order_comment.text or "").strip()
        except Exception:
            order_comment = ""
        ws.merge_cells("A5:B5")
        ws["A5"].value = "Коментарі:"
        ws["A5"].alignment = Alignment(horizontal="left", vertical="top")
        ws.merge_cells("C5:F5")
        ws["C5"].value = order_comment
        ws["C5"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[5].height = 48
        # висота під текст
        try:
            needed_c = estimate_lines(order_comment, chars_per_line=52)
            ws.row_dimensions[5].height = max(18, 15 * needed_c + 2)
        except Exception:
            ws.row_dimensions[5].height = 30 if order_comment else 18

        # ---- Table ----
        header_row = 7
        start_row = header_row + 1
        headers = ["№", "Артикул", "Назва", "Кількість", "Ціна грн.", "Сума грн."]
        hidden_pct_col = 8  # column H (hidden data)
        hidden_pct_header = "PCT"


        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=c, value=h)
            cell.font = bold
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_thin
        ws.row_dimensions[header_row].height = 20

        # Hidden percent column (keeps per-row % for reload, not visible to client)
        ws.cell(row=header_row, column=hidden_pct_col, value=hidden_pct_header).font = bold
        try:
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(hidden_pct_col)].hidden = True
        except Exception:
            pass


        k = self.kurs()
        default_pct = self.CART_DEFAULT_PCT

        total_sum = 0.0
        positions = 0

        # keep stable order: by insertion (cart is dict) -> iterate items
        for i, (article, item) in enumerate(self.cart.items(), start=1):
            p = item.get("product", {}) or {}
            qty = int(item.get("qty", 0) or 0)
            if qty <= 0:
                continue
            positions += 1

            row_pct = int(item.get("pct", default_pct) or default_pct)
            opt_usd = float(p.get("opt_usd", 0.0) or 0.0)
            name = str(p.get("name", "") or "")

            price_base_uah_1 = opt_usd * k
            price_uah_1 = price_base_uah_1 * (1.0 + (row_pct / 100.0))
            sum_uah = price_uah_1 * qty
            total_sum += sum_uah

            r = start_row + (positions - 1)

            c1 = ws.cell(row=r, column=1, value=positions)
            c1.alignment = Alignment(horizontal="center", vertical="top")
            c1.border = border_thin

            c2 = ws.cell(row=r, column=2, value=str(article))
            c2.alignment = Alignment(horizontal="left", vertical="top")
            c2.border = border_thin

            c3 = ws.cell(row=r, column=3, value=name)
            c3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c3.border = border_thin

            c4 = ws.cell(row=r, column=4, value=int(qty))
            c4.alignment = Alignment(horizontal="center", vertical="top")
            c4.border = border_thin

            c5 = ws.cell(row=r, column=5, value=float(money2(price_uah_1)))
            c5.number_format = "# ##0.00"
            c5.alignment = Alignment(horizontal="right", vertical="top")
            c5.border = border_thin

            c6 = ws.cell(row=r, column=6, value=float(money2(sum_uah)))
            c6.number_format = "# ##0.00"
            c6.alignment = Alignment(horizontal="right", vertical="top")
            c6.border = border_thin

            # store pct in hidden column (for later re-open)
            try:
                ws.cell(row=r, column=hidden_pct_col, value=int(row_pct))
            except Exception:
                pass

            needed = estimate_lines(name, chars_per_line=52)
            ws.row_dimensions[r].height = 15 * needed + 2

        last_row = start_row + positions - 1 if positions else start_row

        # Thick line under the table
        thick_row = last_row + 1
        for c in range(1, 7):
            ws.cell(row=thick_row, column=c).border = Border(top=thick)

        # ---- Totals ----
        r1 = thick_row + 1
        ws.merge_cells(f"A{r1}:D{r1}")
        ws["A"+str(r1)].value = f"Всього найменувань {positions}, на суму"
        ws["A"+str(r1)].font = bold
        ws["A"+str(r1)].alignment = Alignment(horizontal="left", vertical="center")

        num_str = f"{money2(total_sum):,.2f}".replace(",", " ").replace(".", ",")
        # Put total in the last (sum) area and align to the right edge
        ws.merge_cells(f"E{r1}:F{r1}")
        ws.cell(row=r1, column=5, value=f"{num_str} грн.").font = bold
        ws.cell(row=r1, column=5).alignment = Alignment(horizontal="right", vertical="center")

        r2 = r1 + 1
        ws.merge_cells(f"A{r2}:F{r2}")
        rate_txt = f"{float(k):,.2f}".replace(",", " ").replace(".", ",")
        ws.cell(row=r2, column=1, value=f"{uah_amount_to_words_uk(total_sum)}, (курс USD {rate_txt})").font = bold
        ws.cell(row=r2, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r2].height = 22

        r3 = r2 + 2
        ws.merge_cells(f"A{r3}:B{r3}")
        ws.cell(row=r3, column=1, value="Менеджер").font = bold
        ws.cell(row=r3, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r3].height = 22

        ws.merge_cells(f"C{r3}:E{r3}")
        sig = ws.cell(row=r3, column=3, value="")
        sig.border = Border(bottom=thin)
        sig.alignment = Alignment(horizontal="center", vertical="center")

        # ---- Print setup ----
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        try:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        except Exception:
            pass
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35
        ws.print_area = f"A1:F{r3}"

        # ---- Meta sheet (hidden) + signature for validation ----
        meta = wb.create_sheet('Meta')
        meta.append(['Signature', ACCOUNT_FILE_MAGIC_VALUE])
        meta.append(['Shop', shop_name])
        meta.append(['SavedAt', dt.isoformat(sep=' ', timespec='seconds')])
        meta.append(['Kurs', k])
        try:
            meta.append(['TopPct', float(self.ids.pct.text)])
        except Exception as e:
            self._notify('Помилка', str(e))
        try:
            meta.append(['OrderComment', order_comment])
        except Exception:
            pass
        try:
            meta.sheet_state = 'hidden'
        except Exception as e:
            self._notify('Помилка', str(e))

        # Limit printing to visible columns A-F (client view)
        try:
            ws.print_area = f"A1:F{ws.max_row}"
        except Exception:
            pass
        wb.save(path)
        return path

    def _validate_account_excel(self, wb, path: str = '') -> None:
        """Validate that the workbook is a рахунок-file created by this app."""
        # 1) Find meta sheet by signature (robust even if sheet was renamed by Excel/openpyxl)
        meta = None
        if wb is None:
            raise ValueError('Не вдалося відкрити книгу Excel.')
        # Prefer sheet named 'Meta' if present, else scan all sheets.
        candidate_names = []
        if 'Meta' in wb.sheetnames:
            candidate_names.append('Meta')
        candidate_names += [n for n in wb.sheetnames if n not in candidate_names]
        for name in candidate_names:
            ws_meta = wb[name]
            try:
                a1 = str(ws_meta['A1'].value).strip()
                b1 = str(ws_meta['B1'].value).strip()
            except Exception:
                continue
            if a1 == 'Signature' and b1 == ACCOUNT_FILE_MAGIC_VALUE:
                meta = ws_meta
                break
        if meta is None:
            raise ValueError('Це не файл рахунку програми (невірний підпис).')
        # store for load step
        self._last_loaded_meta_sheet = meta

        # 2) Detect header row inside sheet "Рахунок" (template has header not in row 1)
        ws = wb['Рахунок'] if 'Рахунок' in wb.sheetnames else wb.active

        def _norm(x):
            return str(x).strip() if x is not None else ""

        old_prefix = ['№', 'Артикул', 'Назва', 'К-сть', '%']
        new_prefix = ['№', 'Артикул', 'Назва', 'Кількість', 'Ціна грн.', 'Сума грн.']

        header_row = None
        header_kind = None  # 'old' or 'new'
        # Scan first 40 rows; header is guaranteed near top in our files
        for r in range(1, 41):
            row_vals = [_norm(c.value) for c in ws[r]]
            # Trim trailing empties
            while row_vals and row_vals[-1] == "":
                row_vals.pop()
            if len(row_vals) >= len(old_prefix) and row_vals[:len(old_prefix)] == old_prefix:
                header_row = r
                header_kind = 'old'
                break
            if len(row_vals) >= len(new_prefix) and row_vals[:len(new_prefix)] == new_prefix:
                header_row = r
                header_kind = 'new'
                break

        if header_row is None:
            raise ValueError('Файл рахунку має невірний формат колонок.')

        # store for loader
        self._last_loaded_header_row = header_row
        self._last_loaded_header_kind = header_kind

    def _load_cart_from_excel(self, path: str) -> None:
        wb = load_workbook(path, data_only=True)
        self._validate_account_excel(wb, path)
        if 'Рахунок' in wb.sheetnames:
            ws = wb['Рахунок']
        else:
            ws = wb.active
        # Restore shop name from Meta if available
        try:
            meta = getattr(self, '_last_loaded_meta_sheet', None) or wb['Meta']
            saved_shop = meta['B2'].value  # row 2: Shop
            if saved_shop is not None and hasattr(self.ids, 'shop'):
                self.ids.shop.text = str(saved_shop)
        except Exception as e:
            self._notify('Помилка', str(e))
        # Restore order comment (Коментарі) from Meta if available
        try:
            oc = ""
            meta_ws = getattr(self, '_last_loaded_meta_sheet', None)
            if meta_ws is not None:
                for r in range(1, 30):
                    k = meta_ws.cell(row=r, column=1).value
                    if str(k).strip() == 'OrderComment':
                        oc = meta_ws.cell(row=r, column=2).value
                        break
            if (oc is None or str(oc).strip() == ''):
                # Fallback: try read from visible sheet row with label "Коментарі:"
                for rr in range(2, 15):
                    if str(ws[f"A{rr}"].value).strip() == "Коментарі:":
                        oc = ws[f"C{rr}"].value
                        break
            if 'order_comment' in self.ids:
                self.ids.order_comment.text = (str(oc) if oc is not None else '')
        except Exception:
            pass
        # Header may be not in row 1 (template). Use detected row from validator.
        header_row = int(getattr(self, '_last_loaded_header_row', 1) or 1)
        header = [c.value for c in ws[header_row]]
        def _norm(x):
            return str(x).strip() if x is not None else ""
        head = [_norm(x) for x in header]

        # hidden pct column (not visible) if present
        pct_col = None
        try:
            for idx, name in enumerate(head, start=1):
                if name.upper() == 'PCT':
                    pct_col = idx
                    break
        except Exception:
            pct_col = None

        # detect format
        is_new = len(head) >= 6 and head[:6] == ['№', 'Артикул', 'Назва', 'Кількість', 'Ціна грн.', 'Сума грн.']


        # pct (markup) restore: prefer Meta TopPct if present
        default_pct = int(getattr(self, 'CART_DEFAULT_PCT', 0) or 0)
        pct_from_meta = None
        try:
            meta = getattr(self, '_last_loaded_meta_sheet', None)
            # row 5: TopPct
            if meta is not None and str(meta['A5'].value).strip() == 'TopPct':
                pct_from_meta = int(float(meta['B5'].value))
        except Exception:
            pct_from_meta = None
        if pct_from_meta is None:
            pct_from_meta = default_pct

        self.cart = {}
        for r in ws.iter_rows(min_row=header_row+1, values_only=True):
            if not r:
                continue

            # --- columns mapping ---
            if is_new:
                # №, Артикул, Назва, Кількість, Ціна грн., Сума грн.
                article = r[1] if len(r) > 1 else None
                name = r[2] if len(r) > 2 else ''
                qty = r[3] if len(r) > 3 else 1
                price_uah_1 = r[4] if len(r) > 4 else None
                row_pct = pct_from_meta
            else:
                # Old: №, Артикул, Назва, К-сть, %, Ціна (опт, грн) ...
                article = r[1] if len(r) > 1 else None
                name = r[2] if len(r) > 2 else ''
                qty = r[3] if len(r) > 3 else 1
                row_pct = int(r[4]) if len(r) > 4 and r[4] is not None else pct_from_meta
                price_uah_1 = r[5] if len(r) > 5 else None

            # Hidden PCT column (preferred when present, keeps per-row % without showing it to client)
            if pct_col is not None and len(r) >= pct_col:
                try:
                    v = r[pct_col - 1]
                    if v is not None and str(v).strip() != '':
                        row_pct = int(float(str(v).replace(',', '.')))
                except Exception:
                    pass
            if article is None:
                continue
            article = str(article).strip()
            if not article:
                continue

            try:
                qty = int(qty)
            except Exception:
                qty = 1
            qty = max(qty, 1)

            name = str(name).strip() if name is not None else ''

            pct = int(row_pct) if row_pct is not None else default_pct

            # Try to find product in current catalog
            p = next((x for x in getattr(self, 'products_all', []) or [] if x.get('article') == article), None)

            if not p:
                # Create minimal product; try to recover opt_usd from saved price (if available)
                opt_usd = 0.0
                try:
                    if price_uah_1 is not None:
                        k = float(self.kurs() or 1.0)
                        price_uah_1 = float(price_uah_1)
                        base_uah = price_uah_1 / (1.0 + (pct / 100.0)) if (1.0 + (pct / 100.0)) else price_uah_1
                        opt_usd = base_uah / k if k else 0.0
                except Exception:
                    opt_usd = 0.0
                p = {'article': article, 'name': name, 'opt_usd': opt_usd, 'retail_usd': 0.0}

            self.cart[article] = {'qty': qty, 'pct': pct, 'product': p}

        self._cart_dirty = False
        self._cart_last_saved_path = str(path)
    def load_cart(self) -> None:
        """Load cart from JSON file."""
        path = self._cart_save_path()
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
        except FileNotFoundError:
            self._notify("Немає файлу", f"Файл не знайдено:\n{path}")
            return
        except Exception as e:
            self._notify("Помилка", f"Не вдалося відкрити файл:\n{e}")
            return
        # map current products by article (if already loaded)
        prod_by_art = {}
        for p in getattr(self, "products_all", []) or []:
            a = p.get("article")
            if a:
                prod_by_art[a] = p
        self.cart = {}
        for it in (payload.get("items") or []):
            art = it.get("article")
            if not art:
                continue
            p = prod_by_art.get(art) or {
                "article": art,
                "name": it.get("name", ""),
                "opt_usd": it.get("opt_usd", 0.0),
                "retail_usd": it.get("retail_usd", 0.0),
            }
            self.cart[art] = {
                "p": p,
                "qty": int(it.get("qty", 1) or 1),
                "pct": float(it.get("pct", 0.0) or 0.0),
            }
        self.refresh_cart()
        self._notify("Відкрито", "Кошик завантажено.")
class TabletApp(App):
    def build(self):
        # Prepare writable data files on Android
        try:
            ensure_runtime_files()
        except Exception:
            pass
        Builder.load_string(KV)
        root = Root()
        # Ещё раз применяем сохранённые значения чуть позже (на всякий случай, если что-то их перезаписывает на старте)
        try:
            Clock.schedule_once(lambda dt: root._apply_ui_state(), 0)
            Clock.schedule_once(lambda dt: root._apply_ui_state(), 0.25)
        except Exception:
            pass
        # Надёжное сохранение курса/% при закрытии окна крестиком (даже если on_stop не сработал как ожидается)
        try:
            from kivy.core.window import Window
            def _req_close(*_args):
                try:
                    if root:
                        root._save_ui_state()
                except Exception as e:
                    try:
                        from kivy.logger import Logger
                        Logger.exception(f"ui_state save on_request_close failed: {e}")
                    except Exception:
                        pass
                return False  # не блокируем закрытие
            Window.bind(on_request_close=_req_close)
        except Exception:
            pass
        return root
if __name__ == '__main__':
    TabletApp().run()